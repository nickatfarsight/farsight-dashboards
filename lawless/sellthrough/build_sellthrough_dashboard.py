#!/usr/bin/env python3
# =============================================================================
# Lawless Beauty — Sell-Through Performance Dashboard Builder
# =============================================================================
# Reads sell-through actuals from the Sales Database and forecasts from the
# Demand Plan, then generates a self-contained interactive HTML dashboard.
#
# Usage:
#   python3 build_sellthrough_dashboard.py
#
# Output:
#   Lawless_SellThrough_Dashboard.html (in the same directory)
#
# Weekly refresh:
#   1. Update "Sales Database with eCommerce.xlsx" in Source Data/
#   2. Re-run this script
#   3. Open the generated HTML in a browser
# =============================================================================

import os
import json
import base64
import re
from datetime import datetime, date, timedelta

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl")
    exit(1)

try:
    import pandas as pd
    import numpy as np
except ImportError:
    print("ERROR: pandas is required. Install with: pip3 install pandas")
    exit(1)

# ---------------------------------------------------------------------------
# Configuration — update these paths when files change
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DASHBOARD_DIR = os.path.join(os.path.expanduser("~"),
    "Library/CloudStorage/OneDrive-Farsight/Sharepoint - Clients/Lawless/"
    "Finance/Claude/Sell Through Dashboard")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "index.html")

SALES_DB_FILE = os.path.join(DASHBOARD_DIR, "Sales Database with eCommerce.xlsx")
DEMAND_PLAN_FILE = os.path.join(DASHBOARD_DIR, "Lawless Demand Plan - Budget.xlsx")
LOCSLS_FILE = os.path.join(DASHBOARD_DIR, "Sephora Sales Location Database.xlsx")

# Retailer sheets to read from Demand Plan (sheet name → Sales DB retailer name)
FORECAST_RETAILER_MAP = {
    "Sephora": "Sephora",
    "Kohls": "Kohls",
    "Ecomm": "eCommerce",
    "ShopBop": "Shopbop",
    "Amazon": "Amazon",
    "GMA": "GMA",
    "QVC": "QVC",
    "Off Price": "Off Price",
    "Other WS": "Other WS",
}

# Month ordering for 4-4-5 calendar
MONTH_ORDER = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}

print("=" * 60)
print("  Lawless Beauty — Sell-Through Performance Dashboard")
print("=" * 60)
print()

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def num(v):
    if v is None:
        return 0
    if isinstance(v, float) and (pd.isna(v) or np.isinf(v)):
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        v = v.strip().replace(",", "").replace("$", "")
        try:
            return float(v)
        except ValueError:
            return 0
    return 0

def safe_div(a, b):
    if b == 0 or b is None:
        return None
    return a / b

def pct_change(ty, ly):
    if ly == 0 or ly is None:
        return None
    return (ty - ly) / abs(ly)

# ---------------------------------------------------------------------------
# 1. READ SALES DATABASE
# ---------------------------------------------------------------------------
print("  [1/5] Reading Sales Database...")
df = pd.read_excel(SALES_DB_FILE, sheet_name="Sales Database", header=0, engine="openpyxl")
print(f"        {len(df):,} rows loaded, {df['Item Code'].nunique()} unique SKUs")

# Normalize column names for easy access
df.columns = df.columns.str.strip()

# Fill NaN in numeric columns
numeric_cols = [
    "TY B&M Sales $", "TY Dotcom Sales $", "TY Total Sales $",
    "TY B&M Sales Units", "TY Dotcom Sales Units", "TY Total Sales Units",
    "LY B&M Sales $", "LY Dotcom Sales $", "LY Total Sales $",
]
for c in numeric_cols:
    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

# Ensure Week and Year are int
df["Week"] = pd.to_numeric(df["Week"], errors="coerce").fillna(0).astype(int)
df["Year"] = pd.to_numeric(df["Year"], errors="coerce").fillna(0).astype(int)

# Detect current period from "Current Week" flag
current_rows = df[df["Current Week"] == "Current"]
if len(current_rows) > 0:
    CURRENT_YEAR = int(current_rows["Year"].iloc[0])
    CURRENT_WEEK = int(current_rows["Week"].iloc[0])
    CURRENT_MONTH = str(current_rows["445 Month"].iloc[0])
    try:
        CURRENT_WEEK_END = pd.to_datetime(current_rows["Week End Date"].iloc[0]).strftime("%m.%d.%Y")
    except Exception:
        CURRENT_WEEK_END = "Unknown"
else:
    # Fallback: use max year/week
    CURRENT_YEAR = int(df["Year"].max())
    max_year_df = df[df["Year"] == CURRENT_YEAR]
    CURRENT_WEEK = int(max_year_df["Week"].max())
    CURRENT_MONTH = str(max_year_df[max_year_df["Week"] == CURRENT_WEEK]["445 Month"].iloc[0])
    CURRENT_WEEK_END = "Unknown"

print(f"        Current period: FY{CURRENT_YEAR} Week {CURRENT_WEEK} ({CURRENT_MONTH}), ending {CURRENT_WEEK_END}")

# Build SKU attribute lookup from the Sales DB (take first occurrence per Item Code)
sku_attrs = (
    df.drop_duplicates(subset="Item Code", keep="first")
    .set_index("Item Code")[
        ["SKU Desc", "Collection", "Category", "Sub Category", "Launch Status",
         "New v Core", "Launch Year", "Sephora Collection", "Season Code"]
    ]
    .to_dict("index")
)

# Build week-end-date lookup for fiscal calendar
week_dates = (
    df[df["Year"] == CURRENT_YEAR]
    .drop_duplicates(subset=["Year", "Week"])
    .set_index("Week")["Week End Date"]
    .to_dict()
)

# Get sorted unique values for filters
all_retailers = sorted(df["Retailer"].dropna().unique().tolist())
all_categories = sorted(df["Category"].dropna().unique().tolist())
all_collections = sorted(df["Collection"].dropna().unique().tolist())
all_years = sorted(df["Year"].unique().tolist())

# ---------------------------------------------------------------------------
# 2. READ DEMAND PLAN (FORECAST)
# ---------------------------------------------------------------------------
print("  [2/5] Reading Demand Plan forecasts...")

forecast_rows = []  # list of dicts: {sku, retailer, week, fcst_units, fcst_dollars}

wb_dp = openpyxl.load_workbook(DEMAND_PLAN_FILE, read_only=True, data_only=True)

# Read Calendar sheet to map week-end dates to fiscal weeks
cal_map = {}  # date → (fy, week_num, month)
ws_cal = wb_dp["Calendar"]
for row in ws_cal.iter_rows(min_row=2, max_row=800, values_only=True):
    fy = row[1]
    week_num = row[2]
    wk_end = row[7]  # Column H = WK END
    month = row[6]   # Column G = MONTH
    if fy is not None and wk_end is not None:
        try:
            fy_int = int(fy)
            wk_int = int(week_num)
            if isinstance(wk_end, datetime):
                cal_map[wk_end.date()] = (fy_int, wk_int, str(month) if month else "")
        except (ValueError, TypeError):
            pass

print(f"        Calendar: {len(cal_map)} week entries loaded")

# Read each retailer forecast sheet (auto-detect layout per sheet)
total_fcst_rows = 0
fcst_warnings = []

for sheet_name, sales_db_retailer in FORECAST_RETAILER_MAP.items():
    if sheet_name not in wb_dp.sheetnames:
        fcst_warnings.append(f"Sheet '{sheet_name}' not found in Demand Plan")
        continue

    ws = wb_dp[sheet_name]

    # --- Auto-detect date header row (scan rows 13-16) ---
    best_row_num = None
    best_row_data = None
    best_date_count = 0
    for scan_row in range(13, 17):
        try:
            row_data = list(ws.iter_rows(min_row=scan_row, max_row=scan_row, values_only=True))[0]
        except (IndexError, StopIteration):
            continue
        n_dates = sum(1 for v in row_data if isinstance(v, datetime) and v.year == 2026)
        if n_dates > best_date_count:
            best_date_count = n_dates
            best_row_num = scan_row
            best_row_data = row_data

    if best_date_count == 0 or best_row_data is None:
        fcst_warnings.append(f"Sheet '{sheet_name}': no 2026 date columns found")
        continue

    date_header_row = best_row_data

    # Extract unique 2026 date columns (first occurrence of each date = unit forecast)
    fcst_cols = []  # [(col_index, date)]
    seen_dates = set()
    for i, v in enumerate(date_header_row):
        if isinstance(v, datetime) and v.year == 2026:
            d = v.date()
            if d not in seen_dates:
                fcst_cols.append((i, d))
                seen_dates.add(d)

    # --- Auto-detect SRP column (find "SRP" label in date header row) ---
    SRP_COL = None
    for i, v in enumerate(date_header_row):
        if isinstance(v, str) and "SRP" in v.upper():
            SRP_COL = i
            break
    if SRP_COL is None:
        SRP_COL = 12  # fallback to legacy position

    # --- Auto-detect NEWNESS row and data start ---
    next_row_num = best_row_num + 1
    try:
        next_row = list(ws.iter_rows(min_row=next_row_num, max_row=next_row_num, values_only=True))[0]
    except (IndexError, StopIteration):
        next_row = ()
    if next_row and next_row[0] and str(next_row[0]).strip().upper() == "NEWNESS":
        data_start_row = best_row_num + 2
    else:
        data_start_row = best_row_num + 1

    # --- Auto-detect SKU column and channel column from first data row ---
    try:
        first_data = list(ws.iter_rows(min_row=data_start_row, max_row=data_start_row, values_only=True))[0]
    except (IndexError, StopIteration):
        fcst_warnings.append(f"Sheet '{sheet_name}': no data rows found")
        continue

    # If col 1 is a channel string (BM/COM), SKU is at col 2
    col1_val = str(first_data[1]).strip().upper() if first_data[1] is not None else ""
    if col1_val in ("BM", "COM", "B&M", "DOTCOM"):
        SKU_COL = 2
        CHAN_COL = 1
    else:
        SKU_COL = 1
        CHAN_COL = None

    # --- Read data rows ---
    sheet_count = 0
    for row in ws.iter_rows(min_row=data_start_row, max_row=900, values_only=True):
        sku_val = row[SKU_COL] if len(row) > SKU_COL else None
        if sku_val is None:
            continue
        try:
            sku_num = int(sku_val)
        except (ValueError, TypeError):
            continue

        srp = num(row[SRP_COL]) if len(row) > SRP_COL else 0
        channel = str(row[CHAN_COL]).strip() if CHAN_COL is not None and row[CHAN_COL] else ""

        for col_idx, col_date in fcst_cols:
            if col_idx < len(row):
                units = num(row[col_idx])
                if units == 0:
                    continue

                # Map date to fiscal week using calendar
                cal_entry = cal_map.get(col_date)
                if cal_entry:
                    fy, wk_num, month = cal_entry
                else:
                    wk_num = fcst_cols.index((col_idx, col_date)) + 1
                    fy = 2026
                    month = ""

                forecast_rows.append({
                    "sku": sku_num,
                    "retailer": sales_db_retailer,
                    "channel": channel,
                    "week": wk_num,
                    "year": fy,
                    "fcst_units": units,
                    "fcst_dollars": units * srp,
                    "srp": srp,
                })
                sheet_count += 1

    total_fcst_rows += sheet_count
    print(f"        {sheet_name}: {sheet_count:,} data pts (row {best_row_num} dates, row {data_start_row}+ data, SKU col {SKU_COL}, SRP col {SRP_COL})")

wb_dp.close()
print(f"        Total forecast rows: {total_fcst_rows:,}")
if fcst_warnings:
    for w in fcst_warnings:
        print(f"        WARNING: {w}")

# Convert forecast to DataFrame and aggregate (sum BM + COM for Sephora)
df_fcst = pd.DataFrame(forecast_rows)
if len(df_fcst) > 0:
    df_fcst_agg = (
        df_fcst.groupby(["sku", "retailer", "year", "week"])
        .agg(fcst_units=("fcst_units", "sum"), fcst_dollars=("fcst_dollars", "sum"))
        .reset_index()
    )
    # Also aggregate across retailers for total forecast
    print(f"        Aggregated to {len(df_fcst_agg):,} SKU × Retailer × Week forecast rows")
else:
    df_fcst_agg = pd.DataFrame(columns=["sku", "retailer", "year", "week", "fcst_units", "fcst_dollars"])
    print("        WARNING: No forecast data loaded")

# ---------------------------------------------------------------------------
# 3. BUILD FORECAST AGGREGATIONS (kept separate from actuals to avoid
#    duplication from Sub Retailer dimension in Sales DB)
# ---------------------------------------------------------------------------
print("  [3/5] Building forecast aggregations...")

# Build week → month mapping — use Sales DB first (complete), then fill from calendar
week_to_month = {}  # (year, week) → month_str
# Primary: Sales DB has definitive 445 Month per year/week
sales_wk_mo = df.drop_duplicates(subset=["Year", "Week"]).set_index(["Year", "Week"])["445 Month"].to_dict()
for (yr, wk), mo in sales_wk_mo.items():
    week_to_month[(int(yr), int(wk))] = str(mo).strip()
# Fallback: Demand Plan calendar (normalize "JAN" → "Jan")
for d, (fy, wk, mo) in cal_map.items():
    key = (int(fy), int(wk))
    if key not in week_to_month and mo:
        week_to_month[key] = str(mo).strip().title()

# Forecast by retailer × week
if len(df_fcst_agg) > 0:
    fcst_ret_week = (
        df_fcst_agg.groupby(["retailer", "year", "week"])
        .agg(fcst_units=("fcst_units", "sum"), fcst_dollars=("fcst_dollars", "sum"))
        .reset_index()
    )
    # "All" retailer totals
    fcst_all_week = (
        df_fcst_agg.groupby(["year", "week"])
        .agg(fcst_units=("fcst_units", "sum"), fcst_dollars=("fcst_dollars", "sum"))
        .reset_index()
    )
    fcst_all_week["retailer"] = "All"
    fcst_week_df = pd.concat([fcst_ret_week, fcst_all_week], ignore_index=True)

    # Forecast by SKU — YTD only (weeks 1 through current week, for fair comparison)
    fcst_by_sku = (
        df_fcst_agg[
            (df_fcst_agg["year"] == CURRENT_YEAR) &
            (df_fcst_agg["week"] <= CURRENT_WEEK)
        ]
        .groupby("sku")
        .agg(fcst_units=("fcst_units", "sum"), fcst_dollars=("fcst_dollars", "sum"))
        .to_dict("index")
    )
    # Forecast by SKU for current month weeks only
    current_month_weeks = set()
    for (yr, wk), mo in week_to_month.items():
        if yr == CURRENT_YEAR and mo == CURRENT_MONTH:
            current_month_weeks.add(wk)
    # Also get month weeks from actuals data (more reliable)
    actuals_month_weeks = set(
        df[(df["Year"] == CURRENT_YEAR) & (df["445 Month"] == CURRENT_MONTH)]["Week"].unique()
    )
    current_month_weeks = current_month_weeks | actuals_month_weeks

    fcst_by_sku_mtd = (
        df_fcst_agg[
            (df_fcst_agg["year"] == CURRENT_YEAR) &
            (df_fcst_agg["week"].isin(current_month_weeks))
        ]
        .groupby("sku")
        .agg(fcst_units=("fcst_units", "sum"), fcst_dollars=("fcst_dollars", "sum"))
        .to_dict("index")
    )

    # Identify future forecast weeks (beyond current actuals)
    actual_max_week = df[df["Year"] == CURRENT_YEAR]["Week"].max() if len(df[df["Year"] == CURRENT_YEAR]) > 0 else 0
    future_fcst = df_fcst_agg[
        (df_fcst_agg["year"] == CURRENT_YEAR) & (df_fcst_agg["week"] > actual_max_week)
    ]
    print(f"        {len(future_fcst):,} forecast-only rows (future weeks {actual_max_week+1}+)")
else:
    fcst_week_df = pd.DataFrame(columns=["retailer", "year", "week", "fcst_units", "fcst_dollars"])
    fcst_by_sku = {}
    fcst_by_sku_mtd = {}
    future_fcst = pd.DataFrame()

# ---------------------------------------------------------------------------
# 4. PRE-AGGREGATE DATA STRUCTURES
# ---------------------------------------------------------------------------
print("  [4/5] Pre-aggregating data for dashboard...")

# Current year filter
df_cy = df[df["Year"] == CURRENT_YEAR].copy()

# --- A. Weekly Totals (Year × Week × Retailer) ---
actuals_cols = {
    "TY Total Sales $": "sum", "TY B&M Sales $": "sum", "TY Dotcom Sales $": "sum",
    "TY Total Sales Units": "sum", "TY B&M Sales Units": "sum", "TY Dotcom Sales Units": "sum",
    "LY Total Sales $": "sum", "LY B&M Sales $": "sum", "LY Dotcom Sales $": "sum",
}
weekly_by_retailer = (
    df.groupby(["Year", "Week", "445 Month", "Retailer"])
    .agg(**{k: (k, v) for k, v in actuals_cols.items()})
    .reset_index()
)
weekly_all = (
    df.groupby(["Year", "Week", "445 Month"])
    .agg(**{k: (k, v) for k, v in actuals_cols.items()})
    .reset_index()
)
weekly_all["Retailer"] = "All"
weekly_totals_df = pd.concat([weekly_by_retailer, weekly_all], ignore_index=True)

# Merge forecast into weekly totals (at the already-aggregated level, no duplication)
if len(fcst_week_df) > 0:
    weekly_totals_df = weekly_totals_df.merge(
        fcst_week_df.rename(columns={"retailer": "Retailer", "year": "Year", "week": "Week"}),
        on=["Year", "Week", "Retailer"],
        how="left",
    )
    weekly_totals_df["fcst_units"] = weekly_totals_df["fcst_units"].fillna(0)
    weekly_totals_df["fcst_dollars"] = weekly_totals_df["fcst_dollars"].fillna(0)
else:
    weekly_totals_df["fcst_units"] = 0
    weekly_totals_df["fcst_dollars"] = 0

# Add week end dates and month numbers
week_end_map = df.drop_duplicates(subset=["Year", "Week"]).set_index(["Year", "Week"])["Week End Date"].to_dict()
weekly_totals_df["week_end"] = weekly_totals_df.apply(
    lambda r: week_end_map.get((r["Year"], r["Week"]), None), axis=1
)
weekly_totals_df["month_num"] = weekly_totals_df["445 Month"].map(MONTH_ORDER).fillna(0).astype(int)
weekly_totals_df["is_current"] = (
    (weekly_totals_df["Year"] == CURRENT_YEAR) & (weekly_totals_df["Week"] == CURRENT_WEEK)
)

# Add future forecast rows for pacing chart
if len(future_fcst) > 0:
    future_agg = (
        future_fcst.groupby(["year", "week"])
        .agg(fcst_units=("fcst_units", "sum"), fcst_dollars=("fcst_dollars", "sum"))
        .reset_index()
        .rename(columns={"year": "Year", "week": "Week"})
    )
    future_agg["Retailer"] = "All"
    future_agg["445 Month"] = ""
    for col in actuals_cols.keys():
        future_agg[col] = 0
    future_agg["week_end"] = None
    future_agg["month_num"] = 0
    future_agg["is_current"] = False
    weekly_totals_df = pd.concat([weekly_totals_df, future_agg], ignore_index=True)

weekly_totals = []
for _, r in weekly_totals_df.iterrows():
    we = r["week_end"]
    we_str = pd.to_datetime(we).strftime("%Y-%m-%d") if pd.notna(we) else ""
    weekly_totals.append({
        "yr": int(r["Year"]), "wk": int(r["Week"]), "mo": str(r["445 Month"]),
        "mo_n": int(r["month_num"]), "ret": r["Retailer"],
        "we": we_str, "cur": bool(r["is_current"]),
        "ty": round(r["TY Total Sales $"], 2),
        "ty_bm": round(r["TY B&M Sales $"], 2),
        "ty_dc": round(r["TY Dotcom Sales $"], 2),
        "ty_u": round(r["TY Total Sales Units"]),
        "ly": round(r["LY Total Sales $"], 2),
        "ly_bm": round(r["LY B&M Sales $"], 2),
        "ly_dc": round(r["LY Dotcom Sales $"], 2),
        "fc_u": round(r["fcst_units"]),
        "fc_d": round(r["fcst_dollars"], 2),
    })

# --- B. Monthly Totals (Year × Month × Retailer) ---
monthly_by_retailer = (
    df.groupby(["Year", "445 Month", "Retailer"])
    .agg(**{k: (k, v) for k, v in actuals_cols.items()})
    .reset_index()
)
monthly_all = (
    df.groupby(["Year", "445 Month"])
    .agg(**{k: (k, v) for k, v in actuals_cols.items()})
    .reset_index()
)
monthly_all["Retailer"] = "All"
monthly_totals_df = pd.concat([monthly_by_retailer, monthly_all], ignore_index=True)
monthly_totals_df["month_num"] = monthly_totals_df["445 Month"].map(MONTH_ORDER).fillna(0).astype(int)

# Merge forecast into monthly totals
if len(fcst_week_df) > 0:
    # Map forecast weeks to months, then aggregate by month × retailer
    fcst_monthly = fcst_week_df.copy()
    fcst_monthly["month"] = fcst_monthly.apply(
        lambda r: week_to_month.get((int(r["year"]), int(r["week"])), ""), axis=1
    )
    fcst_monthly = fcst_monthly[fcst_monthly["month"] != ""]
    fcst_mo_agg = (
        fcst_monthly.groupby(["retailer", "year", "month"])
        .agg(fcst_units=("fcst_units", "sum"), fcst_dollars=("fcst_dollars", "sum"))
        .reset_index()
        .rename(columns={"retailer": "Retailer", "year": "Year", "month": "445 Month"})
    )
    monthly_totals_df = monthly_totals_df.merge(fcst_mo_agg, on=["Year", "445 Month", "Retailer"], how="left")
    monthly_totals_df["fcst_units"] = monthly_totals_df["fcst_units"].fillna(0)
    monthly_totals_df["fcst_dollars"] = monthly_totals_df["fcst_dollars"].fillna(0)
else:
    monthly_totals_df["fcst_units"] = 0
    monthly_totals_df["fcst_dollars"] = 0

monthly_totals = []
for _, r in monthly_totals_df.iterrows():
    monthly_totals.append({
        "yr": int(r["Year"]), "mo": str(r["445 Month"]), "mo_n": int(r["month_num"]),
        "ret": r["Retailer"],
        "ty": round(r["TY Total Sales $"], 2),
        "ty_bm": round(r["TY B&M Sales $"], 2),
        "ty_dc": round(r["TY Dotcom Sales $"], 2),
        "ty_u": round(r["TY Total Sales Units"]),
        "ly": round(r["LY Total Sales $"], 2),
        "fc_u": round(r["fcst_units"]),
        "fc_d": round(r["fcst_dollars"], 2),
    })

# --- C. Category × Month ---
cat_monthly = (
    df.groupby(["Year", "445 Month", "Category"])
    .agg(**{k: (k, v) for k, v in actuals_cols.items()})
    .reset_index()
)
cat_monthly["month_num"] = cat_monthly["445 Month"].map(MONTH_ORDER).fillna(0).astype(int)

# Add forecast by category × month (join forecast SKUs to their categories via sku_attrs)
if len(df_fcst_agg) > 0:
    fcst_cat = df_fcst_agg.copy()
    fcst_cat["Category"] = fcst_cat["sku"].map(
        lambda s: sku_attrs.get(s, {}).get("Category", "Unknown")
    )
    fcst_cat["month"] = fcst_cat.apply(
        lambda r: week_to_month.get((int(r["year"]), int(r["week"])), ""), axis=1
    )
    fcst_cat = fcst_cat[fcst_cat["month"] != ""]
    fcst_cat_agg = (
        fcst_cat.groupby(["year", "month", "Category"])
        .agg(fcst_units=("fcst_units", "sum"), fcst_dollars=("fcst_dollars", "sum"))
        .reset_index()
        .rename(columns={"year": "Year", "month": "445 Month"})
    )
    cat_monthly = cat_monthly.merge(fcst_cat_agg, on=["Year", "445 Month", "Category"], how="left")
    cat_monthly["fcst_units"] = cat_monthly["fcst_units"].fillna(0)
    cat_monthly["fcst_dollars"] = cat_monthly["fcst_dollars"].fillna(0)
else:
    cat_monthly["fcst_units"] = 0
    cat_monthly["fcst_dollars"] = 0

category_monthly = []
for _, r in cat_monthly.iterrows():
    category_monthly.append({
        "yr": int(r["Year"]), "mo": str(r["445 Month"]), "mo_n": int(r["month_num"]),
        "cat": str(r["Category"]),
        "ty": round(r["TY Total Sales $"], 2),
        "ty_u": round(r["TY Total Sales Units"]),
        "ly": round(r["LY Total Sales $"], 2),
        "fc_u": round(r["fcst_units"]),
        "fc_d": round(r["fcst_dollars"], 2),
    })

# --- D. SKU Performance (per-SKU aggregated metrics) ---
l4w_weeks = list(range(max(1, CURRENT_WEEK - 3), CURRENT_WEEK + 1))

# Pre-build forecast lookup by SKU × week (needed for per-period forecast fields)
fcst_sku_week_lookup = {}  # (sku, week) → fcst_dollars
if len(df_fcst_agg) > 0:
    for _, fr in df_fcst_agg[df_fcst_agg["year"] == CURRENT_YEAR].iterrows():
        key = (fr["sku"], int(fr["week"]))
        fcst_sku_week_lookup[key] = fcst_sku_week_lookup.get(key, 0) + fr["fcst_dollars"]

sku_perf_list = []
for item_code, grp in df_cy.groupby("Item Code"):
    # Convert item_code to string for non-numeric codes (e.g., "497CA")
    try:
        ic_display = int(item_code)
    except (ValueError, TypeError):
        ic_display = str(item_code)
    attrs = sku_attrs.get(item_code, {})

    # YTD totals
    ytd = grp.agg({
        "TY Total Sales $": "sum", "TY B&M Sales $": "sum", "TY Dotcom Sales $": "sum",
        "TY Total Sales Units": "sum", "LY Total Sales $": "sum",
    })

    # Forecast from separate aggregation (avoids Sub Retailer duplication)
    sku_fc = fcst_by_sku.get(ic_display, {})
    ytd_fc_units = sku_fc.get("fcst_units", 0)
    ytd_fc_dollars = sku_fc.get("fcst_dollars", 0)

    sku_fc_mtd = fcst_by_sku_mtd.get(ic_display, {})
    mtd_fc_u = sku_fc_mtd.get("fcst_units", 0)
    mtd_fc_d = sku_fc_mtd.get("fcst_dollars", 0)

    # MTD
    mtd_grp = grp[grp["445 Month"] == CURRENT_MONTH]
    mtd_ty = mtd_grp["TY Total Sales $"].sum()
    mtd_ly = mtd_grp["LY Total Sales $"].sum()

    # Current week
    wk_grp = grp[grp["Week"] == CURRENT_WEEK]
    wk_ty = wk_grp["TY Total Sales $"].sum()
    wk_ly = wk_grp["LY Total Sales $"].sum()

    # L4W
    l4w_grp = grp[grp["Week"].isin(l4w_weeks)]
    l4w_ty = l4w_grp["TY Total Sales $"].sum()
    l4w_ly = l4w_grp["LY Total Sales $"].sum()
    l4w_u = l4w_grp["TY Total Sales Units"].sum()

    # Avg weekly
    n_weeks = grp["Week"].nunique()
    avg_wk = ytd["TY Total Sales $"] / n_weeks if n_weeks > 0 else 0

    # Per-retailer breakdown (YTD)
    ret_breakdown = (
        grp.groupby("Retailer")
        .agg(ty=("TY Total Sales $", "sum"), ly=("LY Total Sales $", "sum"))
        .to_dict("index")
    )
    ret_dict = {k: {"ty": round(v["ty"], 2), "ly": round(v["ly"], 2)} for k, v in ret_breakdown.items()}

    # Per-period forecast from SKU × week lookup
    fc_d_wk = fcst_sku_week_lookup.get((ic_display, CURRENT_WEEK), 0)
    fc_d_l4w = sum(fcst_sku_week_lookup.get((ic_display, w), 0) for w in l4w_weeks)

    # Newness flag
    nvc_val = str(attrs.get("New v Core", ""))
    ly_total = ytd["LY Total Sales $"]
    is_new = (nvc_val == "New") or (ly_total == 0)

    # Launch year
    launch_yr_raw = attrs.get("Launch Year", None)
    launch_yr = int(launch_yr_raw) if launch_yr_raw and not pd.isna(launch_yr_raw) else None

    # Risk tier vs forecast
    risk = None
    risk_var_pct = None
    if ytd_fc_dollars > 500:
        ratio = ytd["TY Total Sales $"] / ytd_fc_dollars
        risk_var_pct = round(ratio * 100, 1)
        if ratio < 0.5:
            risk = "critical"
        elif ratio < 0.7:
            risk = "warning"
        elif ratio > 1.3:
            risk = "over"
        else:
            risk = "on_track"

    sku_perf_list.append({
        "ic": ic_display,
        "desc": str(attrs.get("SKU Desc", "")),
        "coll": str(attrs.get("Collection", "")),
        "cat": str(attrs.get("Category", "")),
        "sub_cat": str(attrs.get("Sub Category", "") or ""),
        "nvc": nvc_val,
        "ls": str(attrs.get("Launch Status", "")),
        "lyr": launch_yr,
        "is_new": is_new,
        "risk": risk,
        "risk_pct": risk_var_pct,
        "ytd": round(ytd["TY Total Sales $"], 2),
        "ytd_bm": round(ytd["TY B&M Sales $"], 2),
        "ytd_dc": round(ytd["TY Dotcom Sales $"], 2),
        "ytd_u": round(ytd["TY Total Sales Units"]),
        "ytd_ly": round(ly_total, 2),
        "mtd": round(mtd_ty, 2), "mtd_ly": round(mtd_ly, 2),
        "wk": round(wk_ty, 2), "wk_ly": round(wk_ly, 2),
        "l4w": round(l4w_ty, 2), "l4w_ly": round(l4w_ly, 2), "l4w_u": round(l4w_u),
        "avg_wk": round(avg_wk, 2),
        "fc_u_ytd": round(ytd_fc_units),
        "fc_d_ytd": round(ytd_fc_dollars, 2),
        "fc_u_mtd": round(mtd_fc_u),
        "fc_d_mtd": round(mtd_fc_d, 2),
        "fc_d_wk": round(fc_d_wk, 2),
        "fc_d_l4w": round(fc_d_l4w, 2),
        "rets": ret_dict,
    })

# Sort by YTD descending and add rank
sku_perf_list.sort(key=lambda x: x["ytd"], reverse=True)
for i, s in enumerate(sku_perf_list):
    s["rank"] = i + 1

# --- E. SKU Weekly Series (top 30 by YTD) ---
top_skus = [s["ic"] for s in sku_perf_list[:30]]

sku_weekly_series = []
for ic in top_skus:
    grp = df_cy[df_cy["Item Code"] == ic].sort_values("Week")
    weeks_data = []
    for _, r in grp.groupby("Week").agg({
        "TY Total Sales $": "sum", "LY Total Sales $": "sum",
    }).reset_index().iterrows():
        wk = int(r["Week"])
        weeks_data.append({
            "wk": wk,
            "ty": round(r["TY Total Sales $"], 2),
            "ly": round(r["LY Total Sales $"], 2),
            "fc": round(fcst_sku_week_lookup.get((ic, wk), 0), 2),
        })
    attrs = sku_attrs.get(ic, {})
    sku_weekly_series.append({
        "ic": ic,
        "desc": str(attrs.get("SKU Desc", "")),
        "weeks": weeks_data,
    })

# --- E2. SKU Weekly Detail (for drill-downs) ---
sku_weekly_detail = []
for (item_code, week), grp in df_cy.groupby(["Item Code", "Week"]):
    try:
        ic = int(item_code)
    except (ValueError, TypeError):
        ic = str(item_code)
    wk = int(week)
    sku_weekly_detail.append({
        "ic": ic, "wk": wk,
        "ty": round(grp["TY Total Sales $"].sum(), 2),
        "ly": round(grp["LY Total Sales $"].sum(), 2),
        "ty_u": round(grp["TY Total Sales Units"].sum()),
        "fc_d": round(fcst_sku_week_lookup.get((ic, wk), 0), 2),
    })
print(f"        SKU weekly detail: {len(sku_weekly_detail):,} rows")

# --- E3. New Launches aggregate ---
new_sku_list = [s for s in sku_perf_list if s["is_new"]]
new_sku_ics = set(s["ic"] for s in new_sku_list)
new_with_sales = len([s for s in new_sku_list if s["ytd"] > 0])
new_with_fcst = len([s for s in new_sku_list if s["fc_d_ytd"] > 0])
new_zero_sales = len([s for s in new_sku_list if s["ytd"] == 0])

# Aggregate weekly series for new SKUs
new_agg_weekly = {}
for row in sku_weekly_detail:
    if row["ic"] in new_sku_ics:
        wk = row["wk"]
        if wk not in new_agg_weekly:
            new_agg_weekly[wk] = {"wk": wk, "ty": 0, "ly": 0, "fc_d": 0, "ty_u": 0}
        new_agg_weekly[wk]["ty"] += row["ty"]
        new_agg_weekly[wk]["ly"] += row["ly"]
        new_agg_weekly[wk]["fc_d"] += row["fc_d"]
        new_agg_weekly[wk]["ty_u"] += row["ty_u"]

new_launches = {
    "count": len(new_sku_list),
    "with_sales": new_with_sales,
    "with_fcst": new_with_fcst,
    "zero_sales": new_zero_sales,
    "agg_weekly": sorted(new_agg_weekly.values(), key=lambda x: x["wk"]),
}
print(f"        New launches: {len(new_sku_list)} SKUs ({new_with_sales} with sales, {new_zero_sales} zero)")

# --- F. Retailer Summary ---
# Pre-build forecast lookups by retailer from fcst_week_df (avoids Sub Retailer duplication)
fcst_ret_ytd = {}  # retailer → {fcst_units, fcst_dollars}
fcst_ret_mtd = {}  # retailer → {fcst_dollars}
fcst_ret_wk = {}   # retailer → fcst_dollars for current week
fcst_ret_l4w = {}  # retailer → fcst_dollars for L4W
if len(fcst_week_df) > 0:
    fcst_cy = fcst_week_df[
        (fcst_week_df["year"] == CURRENT_YEAR) &
        (fcst_week_df["week"] <= CURRENT_WEEK)
    ]
    for ret_name in all_retailers + ["All"]:
        ret_fcst = fcst_cy[fcst_cy["retailer"] == ret_name]
        fcst_ret_ytd[ret_name] = {
            "fcst_units": ret_fcst["fcst_units"].sum(),
            "fcst_dollars": ret_fcst["fcst_dollars"].sum(),
        }
        # MTD: filter to current month weeks
        ret_fcst_mtd = ret_fcst[ret_fcst["week"].isin(current_month_weeks)]
        fcst_ret_mtd[ret_name] = {
            "fcst_dollars": ret_fcst_mtd["fcst_dollars"].sum(),
        }
        # Current week forecast
        ret_fcst_wk = ret_fcst[ret_fcst["week"] == CURRENT_WEEK]
        fcst_ret_wk[ret_name] = ret_fcst_wk["fcst_dollars"].sum()
        # L4W forecast
        ret_fcst_l4w = ret_fcst[ret_fcst["week"].isin(l4w_weeks)]
        fcst_ret_l4w[ret_name] = ret_fcst_l4w["fcst_dollars"].sum()

retailer_summary = []
for ret in all_retailers + ["All"]:
    if ret == "All":
        rdf = df_cy
    else:
        rdf = df_cy[df_cy["Retailer"] == ret]

    ytd_ty = rdf["TY Total Sales $"].sum()
    ytd_ly = rdf["LY Total Sales $"].sum()
    ytd_bm = rdf["TY B&M Sales $"].sum()
    ytd_dc = rdf["TY Dotcom Sales $"].sum()
    ytd_u = rdf["TY Total Sales Units"].sum()

    # Forecast from separate aggregation
    ytd_fc_u = fcst_ret_ytd.get(ret, {}).get("fcst_units", 0)
    ytd_fc_d = fcst_ret_ytd.get(ret, {}).get("fcst_dollars", 0)

    mtd_rdf = rdf[rdf["445 Month"] == CURRENT_MONTH]
    mtd_ty = mtd_rdf["TY Total Sales $"].sum()
    mtd_ly = mtd_rdf["LY Total Sales $"].sum()
    mtd_fc_d = fcst_ret_mtd.get(ret, {}).get("fcst_dollars", 0)

    wk_rdf = rdf[rdf["Week"] == CURRENT_WEEK]
    wk_ty = wk_rdf["TY Total Sales $"].sum()
    wk_ly = wk_rdf["LY Total Sales $"].sum()

    l4w_rdf = rdf[rdf["Week"].isin(l4w_weeks)]
    l4w_ty = l4w_rdf["TY Total Sales $"].sum()
    l4w_ly = l4w_rdf["LY Total Sales $"].sum()

    wk_fc_d = fcst_ret_wk.get(ret, 0)
    l4w_fc_d = fcst_ret_l4w.get(ret, 0)

    retailer_summary.append({
        "ret": ret,
        "ytd": round(ytd_ty, 2), "ytd_ly": round(ytd_ly, 2),
        "ytd_bm": round(ytd_bm, 2), "ytd_dc": round(ytd_dc, 2),
        "ytd_u": round(ytd_u),
        "mtd": round(mtd_ty, 2), "mtd_ly": round(mtd_ly, 2),
        "wk": round(wk_ty, 2), "wk_ly": round(wk_ly, 2),
        "l4w": round(l4w_ty, 2), "l4w_ly": round(l4w_ly, 2),
        "fc_u_ytd": round(ytd_fc_u), "fc_d_ytd": round(ytd_fc_d, 2),
        "fc_d_mtd": round(mtd_fc_d, 2),
        "fc_d_wk": round(wk_fc_d, 2),
        "fc_d_l4w": round(l4w_fc_d, 2),
    })

# --- G. Metadata ---
# Determine which retailers have forecast data
forecast_retailers = []
if len(fcst_week_df) > 0:
    forecast_retailers = sorted(fcst_week_df[fcst_week_df["fcst_dollars"] > 0]["retailer"].unique().tolist())
    # Remove "All" from the list if present
    forecast_retailers = [r for r in forecast_retailers if r != "All"]

all_weeks_in_data = sorted(df_cy["Week"].dropna().unique().astype(int).tolist())

metadata = {
    "current_year": CURRENT_YEAR,
    "current_week": CURRENT_WEEK,
    "current_month": CURRENT_MONTH,
    "current_week_end": CURRENT_WEEK_END,
    "retailers": all_retailers,
    "categories": all_categories,
    "collections": all_collections,
    "years": all_years,
    "l4w_weeks": l4w_weeks,
    "build_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
    "data_through_label": f"Data through Week {CURRENT_WEEK} (ending {CURRENT_WEEK_END})",
    "forecast_retailers": forecast_retailers,
    "all_weeks": all_weeks_in_data,
}

# ---------------------------------------------------------------------------
# 4b. READ & AGGREGATE SEPHORA STORE-LEVEL (LOCSLS) DATA
# ---------------------------------------------------------------------------
loc_summary = []
loc_weekly_list = []
loc_meta = {"territories": [], "regions": [], "states": [], "fixtures": [],
            "volumes": [], "loc_current_week": CURRENT_WEEK}

if os.path.exists(LOCSLS_FILE):
    print("  [4b] Reading Sephora store-level data...")
    df_loc = pd.read_excel(LOCSLS_FILE, sheet_name="Loc_Sls_Sephora", engine="openpyxl")
    df_loc["Year"] = pd.to_numeric(df_loc["Year"], errors="coerce").fillna(0).astype(int)
    df_loc["Week"] = pd.to_numeric(df_loc["Week"], errors="coerce").fillna(0).astype(int)
    for c in ["Week End Sales Net $", "Week End Sales Net $ LY",
              "Week End Sales Units", "Week End Inv Retail", "Week End AFS Units",
              "Week End YTD Sales Net $", "Week End YTD Sales Net $ LY"]:
        df_loc[c] = pd.to_numeric(df_loc[c], errors="coerce").fillna(0)
    # Strip whitespace from string columns
    for c in ["Location Desc", "Territory", "Region", "State", "City",
              "Fixture", "Store Volume", "District"]:
        if c in df_loc.columns:
            df_loc[c] = df_loc[c].astype(str).str.strip()

    print(f"        {len(df_loc):,} rows, {df_loc['Location Number'].nunique()} locations")

    # Current year data
    df_loc_cy = df_loc[df_loc["Year"] == CURRENT_YEAR].copy()

    # Current week row for each location (has YTD built in)
    loc_current_wk = df_loc_cy[df_loc_cy["Week"] == CURRENT_WEEK]

    for _, r in loc_current_wk.iterrows():
        loc_num = int(r["Location Number"])
        loc_weeks = df_loc_cy[df_loc_cy["Location Number"] == loc_num]
        n_weeks = loc_weeks["Week"].nunique()

        # Last 5 weeks average
        l5w_start = max(1, CURRENT_WEEK - 4)
        l5w = loc_weeks[loc_weeks["Week"].between(l5w_start, CURRENT_WEEK)]
        l5w_avg = float(l5w["Week End Sales Net $"].mean()) if len(l5w) > 0 else 0

        ytd_sales = float(r["Week End YTD Sales Net $"])
        ytd_avg_wk = ytd_sales / n_weeks if n_weeks > 0 else 0
        wk_sales = float(r["Week End Sales Net $"])
        chg_d = wk_sales - l5w_avg
        chg_p = chg_d / abs(l5w_avg) if l5w_avg != 0 else None

        loc_summary.append({
            "loc": loc_num,
            "name": str(r["Location Desc"]).strip(),
            "terr": str(r.get("Territory", "")).strip(),
            "reg": str(r.get("Region", "")).strip(),
            "st": str(r.get("State", "")).strip(),
            "city": str(r.get("City", "")).strip(),
            "fix": str(r.get("Fixture", "")).strip(),
            "vol": str(r.get("Store Volume", "")).strip(),
            "wk": round(wk_sales, 2),
            "wk_ly": round(float(r["Week End Sales Net $ LY"]), 2),
            "wk_u": int(r["Week End Sales Units"]),
            "ytd": round(ytd_sales, 2),
            "ytd_ly": round(float(r["Week End YTD Sales Net $ LY"]), 2),
            "ytd_avg": round(ytd_avg_wk, 2),
            "l5w_avg": round(l5w_avg, 2),
            "chg_d": round(chg_d, 2),
            "chg_p": round(chg_p, 4) if chg_p is not None else None,
            "inv": round(float(r["Week End Inv Retail"]), 2),
            "afs": int(r["Week End AFS Units"]),
        })

    loc_summary.sort(key=lambda x: x["ytd"], reverse=True)
    print(f"        Location summary: {len(loc_summary)} locations")

    # Per-location weekly series (CY, top 400 by YTD to limit size)
    top_locs = set(d["loc"] for d in loc_summary[:400])
    for loc_num in top_locs:
        loc_wks = df_loc_cy[df_loc_cy["Location Number"] == loc_num].sort_values("Week")
        weeks_data = []
        for _, wr in loc_wks.iterrows():
            weeks_data.append({
                "wk": int(wr["Week"]),
                "ty": round(float(wr["Week End Sales Net $"]), 2),
                "ly": round(float(wr["Week End Sales Net $ LY"]), 2),
                "u": int(wr["Week End Sales Units"]),
            })
        loc_weekly_list.append({"loc": loc_num, "weeks": weeks_data})

    print(f"        Weekly series: {len(loc_weekly_list)} locations × ~{CURRENT_WEEK} weeks")

    # Filter metadata
    loc_meta = {
        "territories": sorted(set(r["terr"] for r in loc_summary if r["terr"] and r["terr"] != "nan")),
        "regions": sorted(set(r["reg"] for r in loc_summary if r["reg"] and r["reg"] != "nan")),
        "states": sorted(set(r["st"] for r in loc_summary if r["st"] and r["st"] != "nan")),
        "fixtures": sorted(set(r["fix"] for r in loc_summary if r["fix"] and r["fix"] != "nan")),
        "volumes": [v for v in ["A++", "A+", "A", "B", "C", "D", "E"]
                    if v in set(r["vol"] for r in loc_summary)],
        "loc_current_week": CURRENT_WEEK,
    }
    del df_loc, df_loc_cy  # free memory
else:
    print("  [4b] Sephora store-level data: file not found, skipping...")

# ---------------------------------------------------------------------------
# 5. BUILD JSON + FONT EMBEDDING + HTML OUTPUT
# ---------------------------------------------------------------------------
print("  [5/6] Building dashboard HTML...")

data_payload = {
    "weekly": weekly_totals,
    "monthly": monthly_totals,
    "cat_monthly": category_monthly,
    "skus": sku_perf_list,
    "sku_series": sku_weekly_series,
    "sku_weekly": sku_weekly_detail,
    "ret_summary": retailer_summary,
    "new_launches": new_launches,
    "meta": metadata,
    "loc_sales": loc_summary,
    "loc_weekly": loc_weekly_list,
    "loc_meta": loc_meta,
}

json_data = json.dumps(data_payload, default=str)
json_size_kb = len(json_data) / 1024
print(f"        JSON data size: {json_size_kb:.1f} KB")

# Embed Halyard Display font
FONT_DIR = os.path.join(SCRIPT_DIR, "..", "..", "Admin", "Halyard Display 2")
font_css_parts = []
font_weights = {
    "fonnts.com-Halyard_Display_Light.otf": ("300", "normal"),
    "fonnts.com-Halyard_Display_Regular.otf": ("400", "normal"),
    "fonnts.com-Halyard_Display_Medium.otf": ("500", "normal"),
    "fonnts.com-Halyard_Display_SemiBold.otf": ("600", "normal"),
    "fonnts.com-Halyard_Display_Bold.otf": ("700", "normal"),
}
for fname, (weight, style) in font_weights.items():
    fpath = os.path.join(FONT_DIR, fname)
    if os.path.exists(fpath):
        with open(fpath, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("ascii")
        font_css_parts.append(
            f"@font-face{{font-family:'Halyard Display';font-weight:{weight};"
            f"font-style:{style};font-display:swap;"
            f"src:url(data:font/otf;base64,{b64}) format('opentype');}}"
        )
FONT_CSS = "\n".join(font_css_parts)

# ---------------------------------------------------------------------------
# HTML TEMPLATE
# ---------------------------------------------------------------------------
HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>LAWLESS — Sell-Through Performance</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
%%FONT_CSS%%

:root {
  --bg:#FFFFFF;--bg-warm:#FBF6F3;--bg-card:#FFFFFF;
  --border:#E8DDD5;--border-light:#F0E8E2;
  --text:#1A1A1A;--text-muted:#7A6E65;--text-light:#A89889;
  --accent:#C9A291;--accent-dark:#B08A78;--accent-light:#E8D5CC;
  --mauve:#C77D8A;--pink:#E84B8A;
  --green:#5CB85C;--green-bg:rgba(92,184,92,.08);
  --yellow:#D4A017;--yellow-bg:rgba(212,160,23,.08);
  --red:#D9534F;--red-bg:rgba(217,83,79,.08);
  --blue:#5B9BD5;--blue-bg:rgba(91,155,213,.08);
  --orange:#FF9800;--orange-bg:rgba(255,152,0,.08);
  --shadow:0 1px 3px rgba(0,0,0,.06),0 1px 2px rgba(0,0,0,.04);
  --shadow-hover:0 4px 12px rgba(0,0,0,.08);
}
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'Inter',sans-serif;background:var(--bg-warm);color:var(--text);line-height:1.5;}

.header{background:linear-gradient(135deg,#FFFFFF 0%,#FBF6F3 100%);border-bottom:1px solid var(--border);padding:20px 32px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100;box-shadow:0 1px 4px rgba(0,0,0,.04);}
.header h1{font-family:'Halyard Display',sans-serif;font-size:26px;font-weight:500;color:var(--text);letter-spacing:6px;text-transform:uppercase;}
.header .subtitle{font-size:13px;color:var(--text-muted);margin-top:2px;font-weight:400;letter-spacing:.5px;}
.date-badge{background:linear-gradient(135deg,var(--accent),var(--accent-dark));color:#fff;padding:8px 20px;border-radius:24px;font-weight:600;font-size:12px;letter-spacing:.3px;max-width:320px;text-align:center;line-height:1.3;}
.fc-scope{font-size:10px;color:var(--text-light);display:flex;align-items:center;gap:4px;margin-top:6px;justify-content:center;}
.fc-dot{width:6px;height:6px;border-radius:50%;background:var(--blue);display:inline-block;}

.global-filters{display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap;}
.global-filters label{font-size:10px;color:var(--text-muted);font-weight:600;text-transform:uppercase;letter-spacing:.5px;display:block;margin-bottom:3px;}
.global-filters select{background:var(--bg-card);border:1px solid var(--border);color:var(--text);padding:7px 12px;border-radius:8px;font-size:12px;cursor:pointer;font-family:'Inter',sans-serif;}
.global-filters select:focus{outline:none;border-color:var(--mauve);box-shadow:0 0 0 2px rgba(199,125,138,.15);}

/* Multi-select dropdown */
.ms-wrap{position:relative;display:inline-block;}
.ms-btn{background:var(--bg-card);border:1px solid var(--border);color:var(--text);padding:7px 12px;border-radius:8px;font-size:12px;cursor:pointer;font-family:'Inter',sans-serif;min-width:150px;text-align:left;display:flex;align-items:center;justify-content:space-between;gap:6px;white-space:nowrap;}
.ms-btn:hover{border-color:var(--accent);}
.ms-btn .arrow{font-size:8px;color:var(--text-light);transition:transform .2s;}
.ms-pop{display:none;position:absolute;top:100%;left:0;min-width:220px;background:var(--bg-card);border:1px solid var(--border);border-radius:10px;box-shadow:var(--shadow-hover);z-index:200;padding:8px 0;margin-top:4px;max-height:320px;overflow-y:auto;}
.ms-pop.open{display:block;}
.ms-pop label{display:flex;align-items:center;gap:8px;padding:6px 14px;font-size:12px;cursor:pointer;transition:background .1s;}
.ms-pop label:hover{background:var(--bg-warm);}
.ms-pop input[type="checkbox"]{accent-color:var(--mauve);width:14px;height:14px;}
.ms-divider{border-top:1px solid var(--border-light);margin:4px 0;}
.ms-compare{padding:8px 14px;display:flex;align-items:center;gap:8px;font-size:11px;color:var(--text-muted);font-weight:500;}
.ms-compare input{accent-color:var(--blue);}

.nav-tabs{display:flex;gap:0;background:var(--bg-card);border-bottom:1px solid var(--border);padding:0 32px;overflow-x:auto;}
.nav-tab{padding:14px 24px;font-size:12px;font-weight:500;color:var(--text-muted);cursor:pointer;border-bottom:2px solid transparent;transition:all .2s;white-space:nowrap;letter-spacing:.3px;text-transform:uppercase;}
.nav-tab:hover{color:var(--text);background:var(--bg-warm);}
.nav-tab.active{color:var(--mauve);border-bottom-color:var(--mauve);font-weight:600;}

.page{display:none;padding:24px 32px;max-width:1600px;margin:0 auto;}
.page.active{display:block;}

.scope-label{font-size:11px;color:var(--text-muted);padding:4px 12px;background:var(--bg-warm);border:1px solid var(--border-light);border-radius:6px;margin-bottom:16px;display:inline-block;}

.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:16px;margin-bottom:24px;}
.kpi-card{background:var(--bg-card);border:1px solid var(--border-light);border-radius:12px;padding:18px;transition:all .2s;box-shadow:var(--shadow);cursor:pointer;}
.kpi-card:hover{box-shadow:var(--shadow-hover);transform:translateY(-1px);border-color:var(--accent-light);}
.kpi-label{font-size:10px;font-weight:600;color:var(--text-muted);text-transform:uppercase;letter-spacing:.8px;margin-bottom:6px;}
.kpi-value{font-size:26px;font-weight:700;color:var(--text);font-family:'Inter',sans-serif;}
.kpi-delta{font-size:11px;font-weight:600;margin-top:4px;}
.kpi-sub{font-size:10px;color:var(--text-light);margin-top:2px;}
.pos{color:var(--green);}.neg{color:var(--red);}

.chart-row{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px;}
.chart-row.full{grid-template-columns:1fr;}
.chart-container{background:var(--bg-card);border:1px solid var(--border-light);border-radius:12px;padding:20px;box-shadow:var(--shadow);}
.chart-title{font-size:13px;font-weight:600;margin-bottom:14px;color:var(--text);letter-spacing:.2px;}
.chart-hint{font-weight:400;font-size:11px;color:var(--text-light);}
canvas{max-height:340px;}

.table-container{background:var(--bg-card);border:1px solid var(--border-light);border-radius:12px;padding:20px;margin-bottom:24px;box-shadow:var(--shadow);overflow-x:auto;}
.table-title{font-size:13px;font-weight:600;margin-bottom:14px;color:var(--text);}
table{width:100%;border-collapse:collapse;font-size:11px;}
th{text-align:left;padding:9px 8px;font-weight:600;color:var(--text-muted);border-bottom:2px solid var(--border);font-size:10px;text-transform:uppercase;letter-spacing:.5px;white-space:nowrap;cursor:pointer;user-select:none;}
th:hover{color:var(--mauve);}
th.r,td.r{text-align:right;}
th .sort-arrow{font-size:9px;margin-left:3px;opacity:.4;}
th.sorted .sort-arrow{opacity:1;color:var(--mauve);}
td{padding:7px 8px;border-bottom:1px solid var(--border-light);white-space:nowrap;}
tr:hover td{background:var(--bg-warm);}
tr.clickable{cursor:pointer;}

.badge{display:inline-block;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:600;letter-spacing:.3px;}
.badge-green{background:var(--green-bg);color:var(--green);}
.badge-red{background:var(--red-bg);color:var(--red);}
.badge-yellow{background:var(--yellow-bg);color:var(--yellow);}
.badge-blue{background:var(--blue-bg);color:var(--blue);}
.badge-mauve{background:rgba(199,125,138,.1);color:var(--mauve);}

.filter-bar{display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap;align-items:center;}
.filter-bar label{font-size:10px;color:var(--text-muted);font-weight:600;text-transform:uppercase;letter-spacing:.5px;}
.filter-bar select,.filter-bar input{background:var(--bg-card);border:1px solid var(--border);color:var(--text);padding:7px 12px;border-radius:8px;font-size:12px;font-family:'Inter',sans-serif;}
.search-box{width:220px;}
.search-box:focus{outline:none;border-color:var(--mauve);box-shadow:0 0 0 2px rgba(199,125,138,.15);}
.search-box::placeholder{color:var(--text-light);}

.period-toggle{display:flex;gap:0;border:1px solid var(--border);border-radius:8px;overflow:hidden;}
.period-btn{padding:7px 14px;font-size:11px;font-weight:500;cursor:pointer;background:var(--bg-card);color:var(--text-muted);border:none;border-right:1px solid var(--border);font-family:'Inter',sans-serif;transition:all .15s;}
.period-btn:last-child{border-right:none;}
.period-btn:hover{background:var(--bg-warm);color:var(--text);}
.period-btn.active{background:var(--mauve);color:#fff;font-weight:600;}

.heatmap-cell{padding:6px 10px;text-align:right;font-weight:600;font-size:11px;}
.hm-green{background:rgba(92,184,92,.15);color:#3a7a3a;}
.hm-light-green{background:rgba(92,184,92,.08);color:#5CB85C;}
.hm-red{background:rgba(217,83,79,.15);color:#b03a37;}
.hm-light-red{background:rgba(217,83,79,.08);color:#D9534F;}
.hm-neutral{background:transparent;}

/* Risk callouts */
.risk-section{background:var(--bg-card);border:1px solid var(--red);border-left:4px solid var(--red);border-radius:12px;padding:20px;margin-bottom:24px;box-shadow:var(--shadow);}
.risk-header{display:flex;align-items:center;gap:8px;margin-bottom:12px;font-size:14px;font-weight:600;color:var(--red);}
.risk-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:8px;}
.risk-item{display:flex;align-items:center;gap:10px;padding:8px 12px;border-radius:8px;font-size:11px;cursor:pointer;transition:background .15s;}
.risk-item:hover{background:var(--bg-warm);}
.risk-item .r-badge{padding:2px 8px;border-radius:4px;font-size:9px;font-weight:700;letter-spacing:.3px;min-width:55px;text-align:center;}
.risk-critical .r-badge{background:var(--red-bg);color:var(--red);}
.risk-warning .r-badge{background:var(--yellow-bg);color:var(--yellow);}
.risk-item .r-name{flex:1;font-weight:500;color:var(--text);}
.risk-item .r-var{font-weight:600;color:var(--text-muted);}

/* Modal */
.modal-overlay{display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(26,26,26,.45);z-index:1000;backdrop-filter:blur(2px);justify-content:center;align-items:center;}
.modal-overlay.active{display:flex;}
.modal{background:var(--bg-card);border-radius:16px;width:90%;max-width:1000px;max-height:85vh;display:flex;flex-direction:column;box-shadow:0 20px 60px rgba(0,0,0,.15);animation:modalIn .2s ease;}
@keyframes modalIn{from{opacity:0;transform:translateY(20px);}to{opacity:1;transform:translateY(0);}}
.modal-header{display:flex;justify-content:space-between;align-items:center;padding:20px 24px;border-bottom:1px solid var(--border-light);}
.modal-header h3{font-size:15px;font-weight:600;color:var(--text);}
.modal-count{font-size:11px;color:var(--text-muted);background:var(--bg-warm);padding:3px 10px;border-radius:12px;margin-left:10px;}
.modal-close{width:32px;height:32px;border-radius:8px;border:none;background:var(--bg-warm);color:var(--text-muted);font-size:18px;cursor:pointer;display:flex;align-items:center;justify-content:center;}
.modal-close:hover{background:var(--red-bg);color:var(--red);}
.modal-body{padding:16px 24px;overflow-y:auto;flex:1;}
.modal-body table{width:100%;border-collapse:collapse;font-size:11px;}
.modal-body th{text-align:left;padding:8px;font-weight:600;color:var(--text-muted);border-bottom:2px solid var(--border);font-size:10px;text-transform:uppercase;letter-spacing:.5px;white-space:nowrap;position:sticky;top:0;background:var(--bg-card);cursor:pointer;}
.modal-body th:hover{color:var(--mauve);}
.modal-body td{padding:6px 8px;border-bottom:1px solid var(--border-light);white-space:nowrap;}
.modal-body tr:hover td{background:var(--bg-warm);}
.modal-body th.r,.modal-body td.r{text-align:right;}
.modal-chart{height:300px;margin-bottom:16px;}

@media(max-width:1000px){
  .chart-row{grid-template-columns:1fr;}
  .kpi-grid{grid-template-columns:repeat(2,1fr);}
  .header{padding:16px;flex-direction:column;gap:8px;text-align:center;}
  .nav-tabs{padding:0 16px;}.page{padding:16px;}
  .global-filters{justify-content:center;}
}
@media(max-width:600px){.kpi-grid{grid-template-columns:1fr;}}

/* Sephora Doors tab */
.door-filters{display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap;align-items:flex-end;}
.door-filters label{font-size:10px;color:var(--text-muted);font-weight:600;text-transform:uppercase;letter-spacing:.5px;display:block;margin-bottom:3px;}
.door-filters select,.door-filters input{background:var(--bg-card);border:1px solid var(--border);color:var(--text);padding:7px 12px;border-radius:8px;font-size:12px;font-family:'Inter',sans-serif;}
.door-search{width:180px;}
.door-search:focus{outline:none;border-color:var(--mauve);box-shadow:0 0 0 2px rgba(199,125,138,.15);}
.door-search::placeholder{color:var(--text-light);}
.region-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:8px;margin-bottom:24px;}
.region-tile{background:var(--bg-card);border:1px solid var(--border-light);border-radius:10px;padding:14px;text-align:center;cursor:pointer;transition:all .2s;}
.region-tile:hover{box-shadow:var(--shadow-hover);transform:translateY(-1px);}
.region-tile.active-filter{border-color:var(--mauve);box-shadow:0 0 0 2px rgba(199,125,138,.15);}
.rt-label{font-size:10px;font-weight:600;color:var(--text-muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.rt-value{font-size:18px;font-weight:700;}
.rt-sub{font-size:10px;color:var(--text-light);margin-top:2px;}

/* Weekly report email-style table */
.email-table{width:100%;border-collapse:collapse;font-size:12px;}
.email-table th{background:var(--accent);color:#fff;padding:10px 8px;font-size:10px;text-transform:uppercase;letter-spacing:.5px;font-weight:600;white-space:nowrap;}
.email-table th.r{text-align:right;}
.email-table td{padding:8px;border-bottom:1px solid var(--border-light);}
.email-table td.r{text-align:right;}
.email-table tr.total-row{font-weight:700;border-top:2px solid var(--border);background:var(--bg-warm);}
.email-table .channel-sub{padding-left:24px;font-weight:400;color:var(--text-muted);}

</style>
</head>
<body>

<div class="header">
  <div>
    <h1>LAWLESS</h1>
    <div class="subtitle">Sell-Through Performance Dashboard</div>
  </div>
  <div style="display:flex;align-items:flex-end;gap:16px;flex-wrap:wrap;">
    <div class="global-filters">
      <div><label>Period</label>
        <div class="period-toggle" id="periodToggle">
          <button class="period-btn" data-p="wk" id="btnWk">Latest Week</button>
          <button class="period-btn" data-p="mtd">MTD</button>
          <button class="period-btn active" data-p="ytd">YTD</button>
          <button class="period-btn" data-p="l4w">L4W</button>
        </div>
      </div>
      <div><label>Retailer</label>
        <div class="ms-wrap" id="msRetailer">
          <button class="ms-btn" type="button" onclick="toggleRetPop(event)"><span id="msRetLabel">All Retailers</span><span class="arrow">&#9660;</span></button>
          <div class="ms-pop" id="msRetPop"></div>
        </div>
      </div>
      <div><label>Category</label><select id="fCategory"><option value="All">All Categories</option></select></div>
      <div><label>Compare vs</label><select id="fCompare"><option value="ly" selected>Last Year</option><option value="fc">Forecast</option><option value="none">None</option></select></div>
    </div>
    <div>
      <div class="date-badge" id="dateBadge">Loading...</div>
      <div class="fc-scope" id="fcScope"></div>
    </div>
  </div>
</div>

<div class="nav-tabs" id="navTabs">
  <div class="nav-tab active" data-page="overview">Overview</div>
  <div class="nav-tab" data-page="sku">SKU Performance</div>
  <div class="nav-tab" data-page="retailer">Retailer Detail</div>
  <div class="nav-tab" data-page="category">Category</div>
  <div class="nav-tab" data-page="forecast">Forecast vs Actuals</div>
  <div class="nav-tab" data-page="launches">New Launches</div>
  <div class="nav-tab" data-page="doors">Sephora Doors</div>
</div>

<div id="page-overview" class="page active"></div>
<div id="page-sku" class="page"></div>
<div id="page-retailer" class="page"></div>
<div id="page-category" class="page"></div>
<div id="page-forecast" class="page"></div>
<div id="page-launches" class="page"></div>
<div id="page-doors" class="page"></div>

<div class="modal-overlay" id="modalOverlay" onclick="if(event.target===this)closeModal()">
  <div class="modal">
    <div class="modal-header">
      <div style="display:flex;align-items:center;"><h3 id="modalTitle"></h3><span class="modal-count" id="modalCount"></span></div>
      <button class="modal-close" onclick="closeModal()">&times;</button>
    </div>
    <div class="modal-body" id="modalBody"></div>
  </div>
</div>

<script>
// =====================================================================
// DATA & CONSTANTS
// =====================================================================
const DATA = %%DATA_PLACEHOLDER%%;
const M = DATA.meta;
const COLORS = ['#C9A291','#C77D8A','#5B9BD5','#5CB85C','#D4A017','#E84B8A','#8B7EC8','#E8A87C','#6BC5B8','#D9534F','#A0C4E8','#F4B183'];
const MO_ORDER = {Jan:1,Feb:2,Mar:3,Apr:4,May:5,Jun:6,Jul:7,Aug:8,Sep:9,Oct:10,Nov:11,Dec:12};
const FCST_RETS = new Set(M.forecast_retailers || []);

Chart.defaults.color='#7A6E65';Chart.defaults.borderColor='#E8DDD5';
Chart.defaults.font.family='Inter,sans-serif';Chart.defaults.font.size=11;
Chart.defaults.plugins.legend.labels.boxWidth=12;

// =====================================================================
// UTILITIES
// =====================================================================
const fmt=n=>n==null||isNaN(n)?'-':'$'+Math.round(n).toLocaleString();
const fmtK=n=>{if(n==null||isNaN(n))return'-';if(Math.abs(n)>=1e6)return'$'+(n/1e6).toFixed(1)+'M';if(Math.abs(n)>=1e3)return'$'+(n/1e3).toFixed(1)+'K';return'$'+Math.round(n).toLocaleString();};
const fmtN=n=>n==null||isNaN(n)?'-':Math.round(n).toLocaleString();
const fmtPct=n=>{if(n==null||isNaN(n))return'-';return(n>=0?'+':'')+(n*100).toFixed(1)+'%';};
const delta=(a,b)=>(!b||b===0)?null:(a-b)/Math.abs(b);
const hasFc=ret=>FCST_RETS.has(ret);

function periodLabel(){
  const p=state.period;
  if(p==='wk')return'Latest Week (W'+M.current_week+')';
  if(p==='mtd')return M.current_month+' MTD';
  if(p==='l4w')return'Last 4 Weeks';
  return'YTD (W1-'+M.current_week+')';
}

function periodVal(r){
  const p=state.period;
  if(p==='wk')return{ty:r.wk,ly:r.wk_ly,fc:r.fc_d_wk||0};
  if(p==='mtd')return{ty:r.mtd,ly:r.mtd_ly,fc:r.fc_d_mtd||0};
  if(p==='l4w')return{ty:r.l4w,ly:r.l4w_ly,fc:r.fc_d_l4w||0};
  return{ty:r.ytd,ly:r.ytd_ly,fc:r.fc_d_ytd||0};
}

function compVal(pv){
  if(state.comparison==='fc')return pv.fc;
  if(state.comparison==='ly')return pv.ly;
  return null;
}
function compLabel(){
  if(state.comparison==='fc')return'vs Forecast';
  if(state.comparison==='ly')return'vs LY';
  return'';
}
function deltaHtml(ty,comp,label){
  if(comp==null||comp===0)return'<span class="kpi-delta" style="color:var(--text-light)">N/A '+(label||compLabel())+'</span>';
  const d=(ty-comp)/Math.abs(comp);const cls=d>=0?'pos':'neg';const arr=d>=0?'&#9650;':'&#9660;';
  return'<span class="kpi-delta '+cls+'">'+arr+' '+Math.abs(d*100).toFixed(1)+'% '+(label||compLabel())+'</span>';
}
function compDeltaHtml(pv){return deltaHtml(pv.ty,compVal(pv));}

// =====================================================================
// STATE
// =====================================================================
let state={period:'ytd',retailers:['All'],compare:false,category:'All',comparison:'ly'};
let charts={};
function dc(id){if(charts[id]){charts[id].destroy();delete charts[id];}}

// =====================================================================
// MULTI-SELECT RETAILER
// =====================================================================
function buildRetailerMultiSelect(){
  const pop=document.getElementById('msRetPop');
  let h='<label><input type="checkbox" value="All" checked onchange="toggleRetAll(this)"> <strong>All Retailers</strong></label>';
  h+='<div class="ms-divider"></div>';
  M.retailers.forEach(r=>{
    const fcBadge=hasFc(r)?'':'<span style="font-size:9px;color:var(--text-light);margin-left:auto;">no forecast</span>';
    h+='<label><input type="checkbox" value="'+r+'" onchange="toggleRetItem()"> '+r+fcBadge+'</label>';
  });
  h+='<div class="ms-divider"></div>';
  h+='<div class="ms-compare"><input type="checkbox" id="retCompare" onchange="state.compare=this.checked;renderActive()"> <span>Compare retailers side-by-side</span></div>';
  pop.innerHTML=h;
}
function toggleRetPop(e){
  e&&e.stopPropagation();
  document.getElementById('msRetPop').classList.toggle('open');
}
function toggleRetAll(cb){
  const boxes=document.querySelectorAll('#msRetPop input[type="checkbox"]:not([value="All"])');
  boxes.forEach(b=>{b.checked=false;});
  cb.checked=true;
  state.retailers=['All'];
  updateRetLabel();renderActive();
}
function toggleRetItem(){
  const boxes=Array.from(document.querySelectorAll('#msRetPop input[type="checkbox"]:not([value="All"])'));
  const allBox=document.querySelector('#msRetPop input[value="All"]');
  const selected=boxes.filter(b=>b.checked).map(b=>b.value);
  if(selected.length===0){allBox.checked=true;state.retailers=['All'];}
  else{allBox.checked=false;state.retailers=selected;}
  updateRetLabel();renderActive();
}
function updateRetLabel(){
  const el=document.getElementById('msRetLabel');
  if(state.retailers[0]==='All')el.textContent='All Retailers';
  else if(state.retailers.length===1)el.textContent=state.retailers[0];
  else el.textContent=state.retailers.length+' Retailers';
}
// Close popover when clicking outside
document.addEventListener('click',e=>{
  const w=document.getElementById('msRetailer');
  if(w&&!w.contains(e.target))document.getElementById('msRetPop').classList.remove('open');
});

// =====================================================================
// DATA HELPERS
// =====================================================================
function effRet(){return state.retailers[0]==='All'?'All':state.retailers.length===1?state.retailers[0]:null;}

function getRetSummary(ret){return DATA.ret_summary.find(r=>r.ret===ret)||DATA.ret_summary.find(r=>r.ret==='All');}

function getAggSummary(){
  // When multiple retailers selected, aggregate their summaries
  if(state.retailers[0]==='All'||state.retailers.length===1)return getRetSummary(effRet()||'All');
  const keys=['ytd','ytd_ly','ytd_bm','ytd_dc','ytd_u','mtd','mtd_ly','wk','wk_ly','l4w','l4w_ly','fc_u_ytd','fc_d_ytd','fc_d_mtd','fc_d_wk','fc_d_l4w'];
  const agg={ret:'Selected'};
  keys.forEach(k=>{agg[k]=0;});
  state.retailers.forEach(r=>{const rs=getRetSummary(r);if(rs)keys.forEach(k=>{agg[k]+=rs[k]||0;});});
  return agg;
}

function fWeekly(yr,ret){
  let d=DATA.weekly.filter(r=>r.yr===yr);
  if(ret&&ret!=='All')d=d.filter(r=>r.ret===ret);
  else d=d.filter(r=>r.ret==='All');
  return d;
}
function fMonthly(yr,ret){
  let d=DATA.monthly.filter(r=>r.yr===yr);
  if(ret&&ret!=='All')d=d.filter(r=>r.ret===ret);
  else d=d.filter(r=>r.ret==='All');
  return d;
}

function filteredSKUs(){
  let s=DATA.skus;
  if(state.category!=='All')s=s.filter(r=>r.cat===state.category);
  if(state.retailers[0]!=='All'){
    s=s.filter(r=>{
      if(!r.rets)return false;
      return state.retailers.some(ret=>r.rets[ret]);
    });
  }
  const coll=document.getElementById('fColl')?.value||'All';
  if(coll!=='All')s=s.filter(r=>r.coll===coll);
  const nvc=document.getElementById('fNVC')?.value||'All';
  if(nvc!=='All')s=s.filter(r=>r.nvc===nvc);
  const q=(document.getElementById('skuSearch')?.value||'').toLowerCase();
  if(q)s=s.filter(r=>r.desc.toLowerCase().includes(q)||String(r.ic).includes(q));
  return s;
}

// Recompute retailer-level data from SKUs when category filter is active
function computeRetFromSKUs(){
  const skus=DATA.skus.filter(s=>state.category!=='All'?s.cat===state.category:true);
  const retData={};
  M.retailers.forEach(r=>{retData[r]={ret:r,ty:0,ly:0};});
  skus.forEach(s=>{
    if(s.rets){Object.entries(s.rets).forEach(([r,v])=>{if(retData[r]){retData[r].ty+=v.ty;retData[r].ly+=v.ly;}});}
  });
  return Object.values(retData).filter(r=>r.ty>0||r.ly>0);
}

function scopeText(){
  let parts=[];
  if(state.retailers[0]==='All')parts.push('All Retailers');
  else parts.push(state.retailers.join(', '));
  if(state.category!=='All')parts.push(state.category);
  else parts.push('All Categories');
  return parts.join(' | ');
}

// =====================================================================
// DRILL-DOWN MODAL
// =====================================================================
let _ddRows=[],_ddCols=[],_ddSortCol=null,_ddSortDir='desc';

function showDrillDown(title,rows,columns){
  _ddRows=rows;_ddCols=columns;_ddSortCol=null;_ddSortDir='desc';
  document.getElementById('modalTitle').textContent=title;
  document.getElementById('modalCount').textContent=rows.length+' items';
  renderDrillBody();
  document.getElementById('modalOverlay').classList.add('active');
}
function renderDrillBody(){
  let sorted=_ddRows.slice();
  if(_ddSortCol){
    sorted.sort((a,b)=>{
      let va=typeof a[_ddSortCol]==='number'?a[_ddSortCol]:(a[_ddSortCol]||'');
      let vb=typeof b[_ddSortCol]==='number'?b[_ddSortCol]:(b[_ddSortCol]||'');
      if(typeof va==='string'){va=va.toLowerCase();vb=(vb||'').toLowerCase();}
      if(va==null)va=_ddSortDir==='asc'?Infinity:-Infinity;
      if(vb==null)vb=_ddSortDir==='asc'?Infinity:-Infinity;
      return _ddSortDir==='asc'?(va<vb?-1:va>vb?1:0):(va>vb?-1:va<vb?1:0);
    });
  }
  let h='<table><thead><tr>';
  _ddCols.forEach(c=>{
    const arrow=_ddSortCol===c.key?(_ddSortDir==='asc'?'&#9650;':'&#9660;'):'&#9650;';
    const cls=(c.align||'')+(_ddSortCol===c.key?' sorted':'');
    h+='<th class="'+cls+'" data-ddcol="'+c.key+'">'+c.label+'<span class="sort-arrow">'+arrow+'</span></th>';
  });
  h+='</tr></thead><tbody>';
  sorted.forEach(r=>{
    h+='<tr>';
    _ddCols.forEach(c=>{
      const v=c.render?c.render(r):(r[c.key]!=null?r[c.key]:'-');
      h+='<td class="'+(c.align||'')+'">'+v+'</td>';
    });
    h+='</tr>';
  });
  h+='</tbody></table>';
  document.getElementById('modalBody').innerHTML=h;
  document.querySelectorAll('#modalBody th[data-ddcol]').forEach(th=>{
    th.addEventListener('click',()=>{
      const col=th.dataset.ddcol;
      if(_ddSortCol===col)_ddSortDir=_ddSortDir==='asc'?'desc':'asc';
      else{_ddSortCol=col;_ddSortDir='desc';}
      renderDrillBody();
    });
  });
}

// Legacy modal for SKU detail with chart
function showModal(title,html){
  document.getElementById('modalTitle').textContent=title;
  document.getElementById('modalCount').textContent='';
  document.getElementById('modalBody').innerHTML=html;
  document.getElementById('modalOverlay').classList.add('active');
}
function closeModal(){document.getElementById('modalOverlay').classList.remove('active');}
document.addEventListener('keydown',e=>{if(e.key==='Escape')closeModal();});

// Standard drill columns for SKU lists
function getDDCols(){return[
  {key:'ic',label:'Item',align:''},{key:'desc',label:'SKU',align:'',render:r=>r.desc.substring(0,35)},
  {key:'cat',label:'Category',align:''},{key:'nvc',label:'N/C',align:''},
  {key:'_pv_ty',label:'Period $',align:'r',render:r=>fmtK(periodVal(r).ty)},
  {key:'_pv_comp',label:state.comparison==='fc'?'Forecast $':'LY $',align:'r',render:r=>{const pv=periodVal(r);return fmtK(compVal(pv));}},
  {key:'_pv_pct',label:'% Chg',align:'r',render:r=>{const pv=periodVal(r);const cv=compVal(pv);const d=delta(pv.ty,cv);return d!=null?'<span class="'+(d>=0?'pos':'neg')+'">'+fmtPct(d)+'</span>':'-';}},
  {key:'ytd',label:'YTD $',align:'r',render:r=>fmtK(r.ytd)},
  {key:'fc_d_ytd',label:'YTD Fcst',align:'r',render:r=>r.fc_d_ytd>0?fmtK(r.fc_d_ytd):'N/A'},
];}

// =====================================================================
// INIT FILTERS
// =====================================================================
const fCat=document.getElementById('fCategory');
M.categories.forEach(c=>{const o=document.createElement('option');o.value=c;o.textContent=c;fCat.appendChild(o);});

document.getElementById('dateBadge').innerHTML=M.data_through_label||('Week '+M.current_week+' &middot; '+M.current_week_end);
document.getElementById('btnWk').textContent='Latest Week (W'+M.current_week+')';

// Forecast scope indicator
const fcRets=M.forecast_retailers||[];
if(fcRets.length>0&&fcRets.length<M.retailers.length){
  document.getElementById('fcScope').innerHTML='<span class="fc-dot"></span> Forecast: '+fcRets.join(', ')+' only';
}

// Period toggle
document.querySelectorAll('.period-btn').forEach(b=>b.addEventListener('click',()=>{
  document.querySelectorAll('.period-btn').forEach(x=>x.classList.remove('active'));
  b.classList.add('active');state.period=b.dataset.p;renderActive();
}));
fCat.onchange=()=>{state.category=fCat.value;renderActive();};
document.getElementById('fCompare').onchange=function(){state.comparison=this.value;renderActive();};

// Tab navigation
document.querySelectorAll('.nav-tab').forEach(t=>t.addEventListener('click',()=>{
  document.querySelectorAll('.nav-tab').forEach(x=>x.classList.remove('active'));
  document.querySelectorAll('.page').forEach(x=>x.classList.remove('active'));
  t.classList.add('active');
  document.getElementById('page-'+t.dataset.page).classList.add('active');
  renderActive();
}));

function activeTab(){return document.querySelector('.nav-tab.active').dataset.page;}
function renderActive(){
  const p=activeTab();
  if(p==='overview')renderOverview();
  else if(p==='sku')renderSKU();
  else if(p==='retailer')renderRetailer();
  else if(p==='category')renderCategory();
  else if(p==='forecast')renderForecast();
  else if(p==='launches')renderNewLaunches();
  else if(p==='doors')renderDoors();
}

// Sort helper
function sortTable(tableId,colIdx,type){
  const tbl=document.getElementById(tableId);if(!tbl)return;
  const tbody=tbl.querySelector('tbody');const rows=Array.from(tbody.querySelectorAll('tr'));
  const th=tbl.querySelectorAll('th')[colIdx];const asc=th.classList.contains('sort-asc');
  tbl.querySelectorAll('th').forEach(h=>{h.classList.remove('sorted','sort-asc','sort-desc');});
  th.classList.add('sorted',asc?'sort-desc':'sort-asc');
  rows.sort((a,b)=>{
    let va=a.children[colIdx].getAttribute('data-v')||a.children[colIdx].textContent;
    let vb=b.children[colIdx].getAttribute('data-v')||b.children[colIdx].textContent;
    if(type==='num'){va=parseFloat(va)||0;vb=parseFloat(vb)||0;}
    else{va=va.toLowerCase();vb=vb.toLowerCase();}
    return asc?(va<vb?-1:va>vb?1:0):(va>vb?-1:va<vb?1:0);
  });
  rows.forEach(r=>tbody.appendChild(r));
}

// Global drill-down helper for KPI cards
function drillKpi(key){
  const skus=filteredSKUs().sort((a,b)=>((b[key]||0)-(a[key]||0)));
  showDrillDown(periodLabel()+' SKU Detail — sorted by '+key,skus,getDDCols());
}

// =====================================================================
// TAB 1: OVERVIEW
// =====================================================================
function buildOverviewPage(){
  document.getElementById('page-overview').innerHTML=`
    <div class="table-container" style="margin-bottom:24px;">
      <div class="table-title" style="display:flex;justify-content:space-between;align-items:center;">
        <span>Sell Through by Customer &mdash; Weekly Report</span>
        <span style="font-size:10px;color:var(--text-light);font-weight:400;">Week ${M.current_week} &middot; ${M.current_month} &middot; FY${M.current_year}</span>
      </div>
      <div id="emailTableWrap"></div>
    </div>
    <div class="kpi-grid" id="ovKpis"></div>
    <div class="chart-row full"><div class="chart-container" style="position:relative;"><div class="chart-title">Weekly Sales Trend <span class="chart-hint">(click point for detail)</span></div><div style="position:relative;height:300px;"><canvas id="ovWeekly"></canvas></div></div></div>
    <div class="chart-row">
      <div class="chart-container"><div class="chart-title">Sales by Retailer <span class="chart-hint">(click bar for detail)</span></div><canvas id="ovRetailer"></canvas></div>
      <div class="chart-container"><div class="chart-title">Monthly Sales <span class="chart-hint">(click bar for detail)</span></div><canvas id="ovMonthly"></canvas></div>
    </div>
    <div class="chart-row">
      <div class="chart-container"><div class="chart-title">Channel Mix &mdash; B&M vs Dotcom (Monthly)</div><canvas id="ovChannel"></canvas></div>
      <div class="chart-container"><div class="chart-title">YoY Growth by Retailer (YTD)</div><canvas id="ovGrowth"></canvas></div>
    </div>
  `;
}

function renderEmailTable(){
  // Build the "Sell Through by Customer" summary table that mirrors the weekly email report
  const rets=DATA.ret_summary.filter(r=>r.ret!=='All').sort((a,b)=>b.ytd-a.ytd);
  const allRow=DATA.ret_summary.find(r=>r.ret==='All')||{};
  let h='<table class="email-table"><thead><tr>';
  h+='<th>Customer</th><th class="r">Week $</th><th class="r">MTD $</th><th class="r">YTD $</th>';
  h+='<th class="r">Wk vs LY</th><th class="r">YTD vs LY</th>';
  h+='<th class="r">Wk vs Fcst</th><th class="r">YTD vs Fcst</th>';
  h+='</tr></thead><tbody>';
  function pctCell(ty,comp){
    if(!comp||comp===0)return'<td class="r" style="color:var(--text-light)">N/A</td>';
    const d=(ty-comp)/Math.abs(comp);const cls=d>=0?'pos':'neg';
    return'<td class="r '+cls+'" style="font-weight:600">'+(d>=0?'+':'')+( d*100).toFixed(1)+'%</td>';
  }
  rets.forEach(r=>{
    const wkD=r.wk_ly?delta(r.wk,r.wk_ly):null;
    const ytdD=r.ytd_ly?delta(r.ytd,r.ytd_ly):null;
    const wkFcD=(r.fc_d_wk&&r.fc_d_wk>0)?delta(r.wk,r.fc_d_wk):null;
    const ytdFcD=(r.fc_d_ytd&&r.fc_d_ytd>0)?delta(r.ytd,r.fc_d_ytd):null;
    h+='<tr>';
    h+='<td style="font-weight:600;">'+r.ret+'</td>';
    h+='<td class="r">'+fmtK(r.wk)+'</td>';
    h+='<td class="r">'+fmtK(r.mtd)+'</td>';
    h+='<td class="r">'+fmtK(r.ytd)+'</td>';
    h+=pctCell(r.wk,r.wk_ly);
    h+=pctCell(r.ytd,r.ytd_ly);
    h+=(r.fc_d_wk>0)?pctCell(r.wk,r.fc_d_wk):'<td class="r" style="color:var(--text-light)">N/A</td>';
    h+=(r.fc_d_ytd>0)?pctCell(r.ytd,r.fc_d_ytd):'<td class="r" style="color:var(--text-light)">N/A</td>';
    h+='</tr>';
  });
  // Total row
  h+='<tr class="total-row">';
  h+='<td>TOTAL</td>';
  h+='<td class="r">'+fmtK(allRow.wk)+'</td>';
  h+='<td class="r">'+fmtK(allRow.mtd)+'</td>';
  h+='<td class="r">'+fmtK(allRow.ytd)+'</td>';
  h+=pctCell(allRow.wk,allRow.wk_ly);
  h+=pctCell(allRow.ytd,allRow.ytd_ly);
  h+=(allRow.fc_d_wk>0)?pctCell(allRow.wk,allRow.fc_d_wk):'<td class="r" style="color:var(--text-light)">N/A</td>';
  h+=(allRow.fc_d_ytd>0)?pctCell(allRow.ytd,allRow.fc_d_ytd):'<td class="r" style="color:var(--text-light)">N/A</td>';
  h+='</tr>';
  h+='</tbody></table>';
  document.getElementById('emailTableWrap').innerHTML=h;
}

function renderOverview(){
  renderEmailTable();
  const rs=getAggSummary();
  const pv=periodVal(rs);
  const cv=compVal(pv);
  const er=effRet()||'All';

  // KPIs — clickable
  let kh='';
  kh+=`<div class="kpi-card" onclick="drillKpi('wk')"><div class="kpi-label">${periodLabel()}</div><div class="kpi-value">${fmtK(pv.ty)}</div>${compDeltaHtml(pv)}<div class="kpi-sub">${scopeText()}</div></div>`;
  kh+=`<div class="kpi-card" onclick="drillKpi('ytd')"><div class="kpi-label">YTD Sales</div><div class="kpi-value">${fmtK(rs.ytd)}</div>${deltaHtml(rs.ytd,rs.ytd_ly,'vs LY')}<div class="kpi-sub">Weeks 1-${M.current_week}</div></div>`;
  kh+=`<div class="kpi-card" onclick="drillKpi('mtd')"><div class="kpi-label">MTD Sales</div><div class="kpi-value">${fmtK(rs.mtd)}</div>${deltaHtml(rs.mtd,rs.mtd_ly,'vs LY')}<div class="kpi-sub">${M.current_month}</div></div>`;
  kh+=`<div class="kpi-card" onclick="drillKpi('l4w')"><div class="kpi-label">L4W Total</div><div class="kpi-value">${fmtK(rs.l4w)}</div>${deltaHtml(rs.l4w,rs.l4w_ly,'vs LY')}<div class="kpi-sub">Last 4 weeks</div></div>`;
  kh+=`<div class="kpi-card" onclick="drillKpi('ytd_u')"><div class="kpi-label">YTD Units</div><div class="kpi-value">${fmtN(rs.ytd_u)}</div><div class="kpi-sub">All channels</div></div>`;
  const fcYtd=rs.fc_d_ytd||0;
  const fcNA=fcYtd<=0;
  kh+=`<div class="kpi-card" onclick="drillKpi('fc_d_ytd')"><div class="kpi-label">YTD Forecast $</div><div class="kpi-value">${fcNA?'N/A':fmtK(fcYtd)}</div>${fcNA?'<span class="kpi-delta" style="color:var(--text-light)">No forecast data</span>':'<span class="kpi-delta '+(rs.ytd>=fcYtd?'pos':'neg')+'">'+(rs.ytd>=fcYtd?'&#9650;':'&#9660;')+' '+Math.abs(delta(rs.ytd,fcYtd)*100).toFixed(1)+'% vs Plan</span>'}<div class="kpi-sub">Demand Plan</div></div>`;
  document.getElementById('ovKpis').innerHTML=kh;

  // --- Weekly trend ---
  dc('ovWeekly');
  const wkTY=fWeekly(M.current_year,er).sort((a,b)=>a.wk-b.wk);
  const wkLY=fWeekly(M.current_year-1,er).sort((a,b)=>a.wk-b.wk);
  const allWks=[...new Set([...wkTY.map(r=>r.wk),...wkLY.map(r=>r.wk)])].sort((a,b)=>a-b);
  const tyMap=Object.fromEntries(wkTY.map(r=>[r.wk,r]));
  const lyMap=Object.fromEntries(wkLY.map(r=>[r.wk,r]));

  const wkDatasets=[
    {label:'TY ('+M.current_year+')',data:allWks.map(w=>(tyMap[w]||{}).ty||null),borderColor:'#C77D8A',backgroundColor:'rgba(199,125,138,.1)',borderWidth:2,fill:true,tension:.3,pointRadius:allWks.map(w=>w===M.current_week?5:2)},
    {label:'LY ('+(M.current_year-1)+')',data:allWks.map(w=>(lyMap[w]||{}).ty||null),borderColor:'#A89889',borderDash:[5,3],borderWidth:1.5,fill:false,tension:.3,pointRadius:0},
  ];
  // Add forecast line if data exists
  const fcData=allWks.map(w=>(tyMap[w]||{}).fc_d||null);
  if(fcData.some(v=>v&&v>0)){
    wkDatasets.push({label:'Forecast',data:fcData,borderColor:'#5B9BD5',borderDash:[4,4],borderWidth:1.5,fill:false,tension:.3,pointRadius:0});
  }

  charts['ovWeekly']=new Chart(document.getElementById('ovWeekly'),{type:'line',data:{
    labels:allWks.map(w=>'W'+w),datasets:wkDatasets
  },options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}},
    onClick:(evt,elements)=>{
      if(!elements.length)return;
      const wk=allWks[elements[0].index];
      const wkSkus=DATA.sku_weekly.filter(r=>r.wk===wk);
      const lookup=Object.fromEntries(DATA.skus.map(s=>[s.ic,s]));
      const rows=wkSkus.map(r=>({...lookup[r.ic],...r,_wk_ty:r.ty,_wk_ly:r.ly,_wk_fc:r.fc_d})).filter(r=>r.desc).sort((a,b)=>(b._wk_ty||0)-(a._wk_ty||0));
      showDrillDown('Week '+wk+' SKU Detail',rows,[
        {key:'ic',label:'Item'},{key:'desc',label:'SKU',render:r=>r.desc.substring(0,35)},{key:'cat',label:'Category'},
        {key:'_wk_ty',label:'TY $',align:'r',render:r=>fmtK(r._wk_ty)},{key:'_wk_ly',label:'LY $',align:'r',render:r=>fmtK(r._wk_ly)},
        {key:'_wk_fc',label:'Forecast $',align:'r',render:r=>r._wk_fc>0?fmtK(r._wk_fc):'N/A'},
        {key:'ty_u',label:'Units',align:'r',render:r=>fmtN(r.ty_u)},
      ]);
    }
  }});

  // --- Retailer bar ---
  dc('ovRetailer');
  // Fix blank chart: when category filter active, recompute from SKUs
  let retChartData;
  if(state.category!=='All'){
    retChartData=computeRetFromSKUs();
  } else {
    retChartData=DATA.ret_summary.filter(r=>r.ret!=='All').map(r=>({ret:r.ret,ty:periodVal(r).ty,ly:periodVal(r).ly,fc:periodVal(r).fc}));
  }
  retChartData.sort((a,b)=>(b.ty||0)-(a.ty||0));
  const retLabels=retChartData.map(r=>r.ret);
  const retDS=[{label:'TY',data:retChartData.map(r=>r.ty||0),backgroundColor:'#C9A291'}];
  if(state.comparison==='ly')retDS.push({label:'LY',data:retChartData.map(r=>r.ly||0),backgroundColor:'#E8DDD5'});
  else if(state.comparison==='fc')retDS.push({label:'Forecast',data:retChartData.map(r=>r.fc||0),backgroundColor:'rgba(91,155,213,.3)'});

  charts['ovRetailer']=new Chart(document.getElementById('ovRetailer'),{type:'bar',data:{
    labels:retLabels,datasets:retDS
  },options:{responsive:true,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}},
    onClick:(evt,elements)=>{
      if(!elements.length)return;
      const retName=retLabels[elements[0].index];
      const skus=DATA.skus.filter(s=>s.rets&&s.rets[retName]).sort((a,b)=>(b.rets[retName]?.ty||0)-(a.rets[retName]?.ty||0));
      showDrillDown(retName+' SKU Detail',skus,[
        {key:'ic',label:'Item'},{key:'desc',label:'SKU',render:r=>r.desc.substring(0,35)},{key:'cat',label:'Category'},
        {key:'_ret_ty',label:'TY $',align:'r',render:r=>fmtK(r.rets[retName]?.ty)},{key:'_ret_ly',label:'LY $',align:'r',render:r=>fmtK(r.rets[retName]?.ly)},
        {key:'ytd',label:'Total YTD',align:'r',render:r=>fmtK(r.ytd)},
      ]);
    }
  }});

  // --- Monthly bar ---
  dc('ovMonthly');
  const moTY=fMonthly(M.current_year,er).sort((a,b)=>a.mo_n-b.mo_n);
  const moLY=fMonthly(M.current_year-1,er).sort((a,b)=>a.mo_n-b.mo_n);
  const allMos=[...new Set([...moTY.map(r=>r.mo),...moLY.map(r=>r.mo)])].sort((a,b)=>(MO_ORDER[a]||0)-(MO_ORDER[b]||0));
  const moTYmap=Object.fromEntries(moTY.map(r=>[r.mo,r]));
  const moLYmap=Object.fromEntries(moLY.map(r=>[r.mo,r]));
  const moDS=[{label:'TY',data:allMos.map(m=>(moTYmap[m]||{}).ty||0),backgroundColor:'#C77D8A'}];
  if(state.comparison==='ly')moDS.push({label:'LY',data:allMos.map(m=>(moLYmap[m]||{}).ty||0),backgroundColor:'#E8DDD5'});
  if(moTY.some(r=>r.fc_d>0))moDS.push({label:'Forecast',data:allMos.map(m=>(moTYmap[m]||{}).fc_d||0),backgroundColor:'rgba(91,155,213,.4)'});

  charts['ovMonthly']=new Chart(document.getElementById('ovMonthly'),{type:'bar',data:{
    labels:allMos,datasets:moDS
  },options:{responsive:true,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}},
    onClick:(evt,elements)=>{
      if(!elements.length)return;
      const mo=allMos[elements[0].index];
      const skus=filteredSKUs().filter(s=>s.mtd>0||s.ytd>0).sort((a,b)=>(b.mtd||0)-(a.mtd||0));
      showDrillDown(mo+' SKU Detail',skus,getDDCols());
    }
  }});

  // --- Channel mix ---
  dc('ovChannel');
  const chData=DATA.monthly.filter(r=>r.yr===M.current_year&&(er==='All'?r.ret==='All':r.ret===er)).sort((a,b)=>a.mo_n-b.mo_n);
  charts['ovChannel']=new Chart(document.getElementById('ovChannel'),{type:'bar',data:{
    labels:chData.map(r=>r.mo),
    datasets:[
      {label:'B&M',data:chData.map(r=>r.ty_bm),backgroundColor:'#C9A291',stack:'s'},
      {label:'Dotcom',data:chData.map(r=>r.ty_dc),backgroundColor:'#5B9BD5',stack:'s'},
    ]
  },options:{responsive:true,plugins:{legend:{position:'top'}},scales:{y:{stacked:true,ticks:{callback:v=>fmtK(v)}},x:{stacked:true}}}});

  // --- YoY Growth ---
  dc('ovGrowth');
  const gRets=DATA.ret_summary.filter(r=>r.ret!=='All'&&r.ytd_ly>0).sort((a,b)=>(b.ytd-b.ytd_ly)/Math.abs(b.ytd_ly)-(a.ytd-a.ytd_ly)/Math.abs(a.ytd_ly));
  const gLabels=gRets.map(r=>r.ret);
  const gData=gRets.map(r=>((r.ytd-r.ytd_ly)/Math.abs(r.ytd_ly))*100);
  const gColors=gData.map(v=>v>=0?'#5CB85C':'#D9534F');
  charts['ovGrowth']=new Chart(document.getElementById('ovGrowth'),{type:'bar',data:{
    labels:gLabels,datasets:[{data:gData,backgroundColor:gColors}]
  },options:{indexAxis:'y',responsive:true,plugins:{legend:{display:false},tooltip:{callbacks:{label:ctx=>ctx.raw.toFixed(1)+'%'}}},scales:{x:{ticks:{callback:v=>v+'%'}}}}});
}

// =====================================================================
// TAB 2: SKU PERFORMANCE
// =====================================================================
function buildSKUPage(){
  document.getElementById('page-sku').innerHTML=`
    <div class="filter-bar">
      <div><label>Collection</label><br><select id="fColl"><option value="All">All Collections</option></select></div>
      <div><label>New / Core</label><br><select id="fNVC"><option value="All">All</option><option value="New">New</option><option value="Core">Core</option></select></div>
      <div><label>Search</label><br><input class="search-box" id="skuSearch" placeholder="Search SKU..."></div>
    </div>
    <div class="kpi-grid" id="skuKpis"></div>
    <div class="chart-row">
      <div class="chart-container"><div class="chart-title">Top 20 SKUs &mdash; ${periodLabel()} <span class="chart-hint">(click bar for detail)</span></div><canvas id="skuTop20"></canvas></div>
      <div class="chart-container"><div class="chart-title">New vs Core Performance <span class="chart-hint">(click segment)</span></div><canvas id="skuNVC"></canvas></div>
    </div>
    <div class="table-container"><div class="table-title">SKU Leaderboard</div><div id="skuTableWrap"></div></div>
  `;
  M.collections.forEach(c=>{const o=document.createElement('option');o.value=c;o.textContent=c;document.getElementById('fColl').appendChild(o);});
  document.getElementById('fColl').onchange=()=>renderSKU();
  document.getElementById('fNVC').onchange=()=>renderSKU();
  document.getElementById('skuSearch').oninput=()=>renderSKU();
}

function renderSKU(){
  const skus=filteredSKUs();
  const totalPV=skus.reduce((a,s)=>a+periodVal(s).ty,0);
  const totalYTD=skus.reduce((a,s)=>a+s.ytd,0);
  const top10sum=skus.slice(0,Math.min(10,skus.length)).reduce((a,s)=>a+s.ytd,0);
  const newSum=skus.filter(s=>s.nvc==='New').reduce((a,s)=>a+s.ytd,0);
  const coreSum=skus.filter(s=>s.nvc==='Core').reduce((a,s)=>a+s.ytd,0);
  const activeSKUs=skus.filter(s=>s.ytd>0).length;
  const totalFcst=skus.reduce((a,s)=>a+(s.fc_d_ytd||0),0);
  const fcAcc=totalFcst>0?(totalYTD/totalFcst*100).toFixed(1)+'%':'N/A';

  let kh='';
  kh+=`<div class="kpi-card" onclick="drillKpi('ytd')"><div class="kpi-label">Active SKUs</div><div class="kpi-value">${fmtN(activeSKUs)}</div><div class="kpi-sub">With YTD sales > $0</div></div>`;
  kh+=`<div class="kpi-card" onclick="drillKpi('ytd')"><div class="kpi-label">Top 10 Concentration</div><div class="kpi-value">${totalYTD>0?(top10sum/totalYTD*100).toFixed(1)+'%':'-'}</div><div class="kpi-sub">Top 10 SKUs % of total</div></div>`;
  kh+=`<div class="kpi-card" onclick="drillKpi('ytd')"><div class="kpi-label">New SKU Contribution</div><div class="kpi-value">${totalYTD>0?(newSum/totalYTD*100).toFixed(1)+'%':'-'}</div><div class="kpi-sub">${fmtK(newSum)} of ${fmtK(totalYTD)}</div></div>`;
  kh+=`<div class="kpi-card" onclick="drillKpi('ytd')"><div class="kpi-label">Avg $ / SKU (YTD)</div><div class="kpi-value">${activeSKUs>0?fmtK(totalYTD/activeSKUs):'-'}</div><div class="kpi-sub">${fmtN(activeSKUs)} active</div></div>`;
  kh+=`<div class="kpi-card" onclick="drillKpi('fc_d_ytd')"><div class="kpi-label">Forecast Accuracy</div><div class="kpi-value">${fcAcc}</div><div class="kpi-sub">${fmtK(totalYTD)} actual / ${fmtK(totalFcst)} plan</div></div>`;
  document.getElementById('skuKpis').innerHTML=kh;

  // Top 20 bar with forecast
  dc('skuTop20');
  const top=skus.slice(0,20);
  const topDS=[
    {label:'TY',data:top.map(s=>periodVal(s).ty),backgroundColor:'#C77D8A'},
  ];
  if(state.comparison==='ly')topDS.push({label:'LY',data:top.map(s=>periodVal(s).ly),backgroundColor:'#E8DDD5'});
  if(top.some(s=>periodVal(s).fc>0))topDS.push({label:'Forecast',data:top.map(s=>periodVal(s).fc),backgroundColor:'rgba(91,155,213,.35)'});

  charts['skuTop20']=new Chart(document.getElementById('skuTop20'),{type:'bar',data:{
    labels:top.map(s=>s.desc.substring(0,30)),datasets:topDS
  },options:{indexAxis:'y',responsive:true,plugins:{legend:{position:'top'}},scales:{x:{ticks:{callback:v=>fmtK(v)}}},
    onClick:(evt,elements)=>{if(elements.length)showSKUDetail(top[elements[0].index].ic);}
  }});

  // New vs Core doughnut
  dc('skuNVC');
  charts['skuNVC']=new Chart(document.getElementById('skuNVC'),{type:'doughnut',data:{
    labels:['New','Core'],datasets:[{data:[newSum,coreSum],backgroundColor:['#C77D8A','#C9A291']}]
  },options:{responsive:true,plugins:{legend:{position:'bottom'},tooltip:{callbacks:{label:ctx=>fmtK(ctx.raw)+' ('+((ctx.raw/(newSum+coreSum||1))*100).toFixed(1)+'%)'}}},
    onClick:(evt,elements)=>{
      if(!elements.length)return;
      const seg=elements[0].index===0?'New':'Core';
      const segSkus=skus.filter(s=>s.nvc===seg).sort((a,b)=>b.ytd-a.ytd);
      showDrillDown(seg+' SKUs',segSkus,getDDCols());
    }
  }});

  // Table
  const getPV=(s)=>periodVal(s);
  let th='<table id="skuTable"><thead><tr>';
  const cols=[
    {l:'#',k:'rank',t:'num',a:'r'},{l:'Item',k:'ic',t:'num',a:''},{l:'SKU',k:'desc',t:'str',a:''},
    {l:'Category',k:'cat',t:'str',a:''},{l:'N/C',k:'nvc',t:'str',a:''},
    {l:'Period $',k:'pv_ty',t:'num',a:'r'},{l:state.comparison==='fc'?'Fcst $':'LY $',k:'pv_comp',t:'num',a:'r'},{l:'% Chg',k:'pv_pct',t:'num',a:'r'},
    {l:'YTD $',k:'ytd',t:'num',a:'r'},{l:'Fcst YTD',k:'fc_d_ytd',t:'num',a:'r'},{l:'L4W $',k:'l4w',t:'num',a:'r'},
  ];
  cols.forEach((c,i)=>{th+=`<th class="${c.a}" onclick="sortTable('skuTable',${i},'${c.t}')">${c.l}<span class="sort-arrow">&#9650;</span></th>`;});
  th+='</tr></thead><tbody>';
  skus.forEach(s=>{
    const pv=getPV(s);const cv=compVal(pv);const pct=delta(pv.ty,cv);
    const pctCls=pct==null?'':(pct>=0?'pos':'neg');
    th+=`<tr class="clickable" onclick="showSKUDetail(${s.ic})">`;
    th+=`<td class="r" data-v="${s.rank}">${s.rank}</td>`;
    th+=`<td data-v="${s.ic}">${s.ic}</td>`;
    th+=`<td>${s.desc.substring(0,35)}</td>`;
    th+=`<td>${s.cat}</td><td>${s.nvc}</td>`;
    th+=`<td class="r" data-v="${pv.ty}">${fmtK(pv.ty)}</td>`;
    th+=`<td class="r" data-v="${cv||0}">${cv!=null?fmtK(cv):'N/A'}</td>`;
    th+=`<td class="r ${pctCls}" data-v="${pct||0}">${fmtPct(pct)}</td>`;
    th+=`<td class="r" data-v="${s.ytd}">${fmtK(s.ytd)}</td>`;
    th+=`<td class="r" data-v="${s.fc_d_ytd}">${s.fc_d_ytd>0?fmtK(s.fc_d_ytd):'N/A'}</td>`;
    th+=`<td class="r" data-v="${s.l4w}">${fmtK(s.l4w)}</td>`;
    th+='</tr>';
  });
  th+='</tbody></table>';
  document.getElementById('skuTableWrap').innerHTML=th;
}

function showSKUDetail(ic){
  const series=DATA.sku_series.find(s=>s.ic===ic);
  const sku=DATA.skus.find(s=>s.ic===ic);
  if(!sku)return;
  let h=`<div class="modal-chart"><canvas id="modalChart"></canvas></div>`;
  h+=`<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px;margin:12px 0 16px;">`;
  h+=`<div class="kpi-card" style="padding:12px;cursor:default;"><div class="kpi-label">YTD $</div><div class="kpi-value" style="font-size:18px;">${fmtK(sku.ytd)}</div></div>`;
  h+=`<div class="kpi-card" style="padding:12px;cursor:default;"><div class="kpi-label">YTD Forecast</div><div class="kpi-value" style="font-size:18px;">${sku.fc_d_ytd>0?fmtK(sku.fc_d_ytd):'N/A'}</div></div>`;
  h+=`<div class="kpi-card" style="padding:12px;cursor:default;"><div class="kpi-label">Avg Weekly $</div><div class="kpi-value" style="font-size:18px;">${fmtK(sku.avg_wk)}</div></div>`;
  if(sku.risk){const rc=sku.risk==='critical'?'var(--red)':sku.risk==='warning'?'var(--yellow)':sku.risk==='over'?'var(--blue)':'var(--green)';
    h+=`<div class="kpi-card" style="padding:12px;cursor:default;border-color:${rc}"><div class="kpi-label">Risk Status</div><div class="kpi-value" style="font-size:16px;color:${rc}">${sku.risk.toUpperCase()}</div><div class="kpi-sub">${sku.risk_pct!=null?sku.risk_pct+'% of plan':''}</div></div>`;}
  h+=`</div>`;
  h+=`<h4 style="margin:12px 0 8px;font-size:13px;">Retailer Breakdown (YTD)</h4>`;
  h+=`<table><thead><tr><th>Retailer</th><th class="r">TY $</th><th class="r">LY $</th><th class="r">% Chg</th></tr></thead><tbody>`;
  if(sku.rets){
    Object.entries(sku.rets).sort((a,b)=>b[1].ty-a[1].ty).forEach(([r,v])=>{
      const d=delta(v.ty,v.ly);const cls=d==null?'':(d>=0?'pos':'neg');
      h+=`<tr><td>${r}</td><td class="r">${fmtK(v.ty)}</td><td class="r">${fmtK(v.ly)}</td><td class="r ${cls}">${fmtPct(d)}</td></tr>`;
    });
  }
  h+='</tbody></table>';
  showModal(sku.desc,h);
  document.getElementById('modalCount').textContent=sku.cat+' | '+sku.nvc+(sku.is_new?' | NEW':'');
  if(series){
    dc('modalChart');
    charts['modalChart']=new Chart(document.getElementById('modalChart'),{type:'line',data:{
      labels:series.weeks.map(w=>'W'+w.wk),
      datasets:[
        {label:'TY $',data:series.weeks.map(w=>w.ty),borderColor:'#C77D8A',borderWidth:2,tension:.3,fill:false},
        {label:'LY $',data:series.weeks.map(w=>w.ly),borderColor:'#A89889',borderDash:[5,3],borderWidth:1.5,tension:.3,fill:false},
        {label:'Forecast $',data:series.weeks.map(w=>w.fc),borderColor:'#5B9BD5',borderDash:[3,3],borderWidth:1.5,tension:.3,fill:false},
      ]
    },options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}}}});
  }
}

// =====================================================================
// TAB 3: RETAILER DETAIL
// =====================================================================
function buildRetailerPage(){
  document.getElementById('page-retailer').innerHTML=`
    <div class="scope-label" id="retScopeLabel">Showing: All Retailers</div>
    <div class="kpi-grid" id="retKpis"></div>
    <div class="chart-row full"><div class="chart-container" style="position:relative;"><div class="chart-title">Weekly Trend <span class="chart-hint">(click point for detail)</span></div><div style="position:relative;height:300px;"><canvas id="retWeekly"></canvas></div></div></div>
    <div class="chart-row">
      <div class="chart-container"><div class="chart-title">Top 10 SKUs <span class="chart-hint">(click bar for detail)</span></div><canvas id="retTopSKU"></canvas></div>
      <div class="chart-container"><div class="chart-title">Monthly Sales — B&M vs Dotcom</div><canvas id="retCatMo"></canvas></div>
    </div>
  `;
}

function renderRetailer(){
  // Use global retailer filter — no per-tab dropdown
  const rets=state.retailers[0]==='All'?M.retailers:state.retailers;
  const isCompare=state.compare&&rets.length>1;
  const firstRet=rets[0]||'Sephora';
  const rs=isCompare?getRetSummary(firstRet):getAggSummary();

  document.getElementById('retScopeLabel').textContent='Showing: '+(state.retailers[0]==='All'?'All Retailers':state.retailers.join(', '));

  // KPIs
  let kh='';
  const pv=periodVal(rs);
  kh+=`<div class="kpi-card"><div class="kpi-label">${periodLabel()}</div><div class="kpi-value">${fmtK(pv.ty)}</div>${compDeltaHtml(pv)}</div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">YTD $</div><div class="kpi-value">${fmtK(rs.ytd)}</div>${deltaHtml(rs.ytd,rs.ytd_ly,'vs LY')}</div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">MTD $</div><div class="kpi-value">${fmtK(rs.mtd)}</div>${deltaHtml(rs.mtd,rs.mtd_ly,'vs LY')}</div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">B&M / Dotcom</div><div class="kpi-value">${rs.ytd_bm+rs.ytd_dc>0?Math.round(rs.ytd_bm/(rs.ytd_bm+rs.ytd_dc)*100)+'% / '+Math.round(rs.ytd_dc/(rs.ytd_bm+rs.ytd_dc)*100)+'%':'-'}</div><div class="kpi-sub">${fmtK(rs.ytd_bm)} B&M &middot; ${fmtK(rs.ytd_dc)} Dotcom</div></div>`;
  const fcYtd=rs.fc_d_ytd||0;
  kh+=`<div class="kpi-card"><div class="kpi-label">YTD Forecast</div><div class="kpi-value">${fcYtd>0?fmtK(fcYtd):'N/A'}</div>${fcYtd>0?deltaHtml(rs.ytd,fcYtd,'vs Plan'):'<span class="kpi-delta" style="color:var(--text-light)">No forecast</span>'}</div>`;
  document.getElementById('retKpis').innerHTML=kh;

  // Weekly trend — compare mode or single
  dc('retWeekly');
  if(isCompare){
    // Multiple lines per retailer
    const allWksSet=new Set();
    const retDatasets=[];
    rets.forEach((ret,i)=>{
      const wk=DATA.weekly.filter(r=>r.yr===M.current_year&&r.ret===ret).sort((a,b)=>a.wk-b.wk);
      wk.forEach(r=>allWksSet.add(r.wk));
      const wkMap=Object.fromEntries(wk.map(r=>[r.wk,r.ty]));
      retDatasets.push({label:ret,data:[],borderColor:COLORS[i%COLORS.length],borderWidth:2,tension:.3,_wkMap:wkMap});
    });
    const allWks=[...allWksSet].sort((a,b)=>a-b);
    retDatasets.forEach(ds=>{ds.data=allWks.map(w=>ds._wkMap[w]||0);delete ds._wkMap;});
    charts['retWeekly']=new Chart(document.getElementById('retWeekly'),{type:'line',data:{
      labels:allWks.map(w=>'W'+w),datasets:retDatasets
    },options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}}}});
  } else {
    const ret=effRet()||firstRet;
    const wkTY=DATA.weekly.filter(r=>r.yr===M.current_year&&r.ret===ret).sort((a,b)=>a.wk-b.wk);
    const wkLY=DATA.weekly.filter(r=>r.yr===M.current_year-1&&r.ret===ret).sort((a,b)=>a.wk-b.wk);
    const allWks=[...new Set([...wkTY.map(r=>r.wk),...wkLY.map(r=>r.wk)])].sort((a,b)=>a-b);
    const tyM=Object.fromEntries(wkTY.map(r=>[r.wk,r]));
    const lyM=Object.fromEntries(wkLY.map(r=>[r.wk,r]));
    const datasets=[
      {label:'TY Total',data:allWks.map(w=>(tyM[w]||{}).ty||0),borderColor:'#C77D8A',borderWidth:2,tension:.3},
      {label:'LY Total',data:allWks.map(w=>(lyM[w]||{}).ty||0),borderColor:'#A89889',borderDash:[5,3],borderWidth:1.5,tension:.3},
    ];
    // Forecast line
    const fcLine=allWks.map(w=>(tyM[w]||{}).fc_d||null);
    if(fcLine.some(v=>v&&v>0)){
      datasets.push({label:'Forecast',data:fcLine,borderColor:'#5B9BD5',borderDash:[4,4],borderWidth:1.5,tension:.3,pointRadius:0});
    }
    charts['retWeekly']=new Chart(document.getElementById('retWeekly'),{type:'line',data:{
      labels:allWks.map(w=>'W'+w),datasets
    },options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}},
      onClick:(evt,elements)=>{
        if(!elements.length)return;
        const wk=allWks[elements[0].index];
        const wkSkus=DATA.sku_weekly.filter(r=>r.wk===wk);
        const lookup=Object.fromEntries(DATA.skus.map(s=>[s.ic,s]));
        const rows=wkSkus.map(r=>({...lookup[r.ic],...r})).filter(r=>r.desc).sort((a,b)=>(b.ty||0)-(a.ty||0));
        showDrillDown('Week '+wk+' Detail',rows,[
          {key:'ic',label:'Item'},{key:'desc',label:'SKU',render:r=>r.desc.substring(0,35)},
          {key:'ty',label:'TY $',align:'r',render:r=>fmtK(r.ty)},{key:'ly',label:'LY $',align:'r',render:r=>fmtK(r.ly)},
          {key:'fc_d',label:'Forecast',align:'r',render:r=>r.fc_d>0?fmtK(r.fc_d):'N/A'},
        ]);
      }
    }});
  }

  // Top 10 SKUs
  dc('retTopSKU');
  const retForSKU=effRet()||firstRet;
  const skusByRet=DATA.skus.filter(s=>s.rets&&s.rets[retForSKU]).sort((a,b)=>(b.rets[retForSKU]?.ty||0)-(a.rets[retForSKU]?.ty||0)).slice(0,10);
  charts['retTopSKU']=new Chart(document.getElementById('retTopSKU'),{type:'bar',data:{
    labels:skusByRet.map(s=>s.desc.substring(0,25)),
    datasets:[
      {label:'TY YTD',data:skusByRet.map(s=>s.rets[retForSKU]?.ty||0),backgroundColor:'#C77D8A'},
      {label:'LY YTD',data:skusByRet.map(s=>s.rets[retForSKU]?.ly||0),backgroundColor:'#E8DDD5'},
    ]
  },options:{indexAxis:'y',responsive:true,plugins:{legend:{position:'top'}},scales:{x:{ticks:{callback:v=>fmtK(v)}}},
    onClick:(evt,elements)=>{if(elements.length)showSKUDetail(skusByRet[elements[0].index].ic);}
  }});

  // Monthly B&M vs Dotcom
  dc('retCatMo');
  const ret2=effRet()||firstRet;
  const retMo=DATA.monthly.filter(r=>r.yr===M.current_year&&r.ret===ret2).sort((a,b)=>a.mo_n-b.mo_n);
  charts['retCatMo']=new Chart(document.getElementById('retCatMo'),{type:'bar',data:{
    labels:retMo.map(r=>r.mo),
    datasets:[
      {label:'B&M',data:retMo.map(r=>r.ty_bm),backgroundColor:'#C9A291',stack:'s'},
      {label:'Dotcom',data:retMo.map(r=>r.ty_dc),backgroundColor:'#5B9BD5',stack:'s'},
    ]
  },options:{responsive:true,plugins:{legend:{position:'top'}},scales:{y:{stacked:true,ticks:{callback:v=>fmtK(v)}},x:{stacked:true}}}});
}

// =====================================================================
// TAB 4: CATEGORY
// =====================================================================
function buildCategoryPage(){
  document.getElementById('page-category').innerHTML=`
    <div class="scope-label" id="catScopeLabel">Showing: All Retailers</div>
    <div class="table-container"><div class="table-title">Category Performance Summary</div><div id="catHeatmap"></div></div>
    <div class="chart-row full"><div class="chart-container" style="position:relative;"><div class="chart-title">Category Monthly Trend (TY)</div><div style="position:relative;height:300px;"><canvas id="catWeekly"></canvas></div></div></div>
    <div class="chart-row">
      <div class="chart-container"><div class="chart-title">Category Share of YTD Sales <span class="chart-hint">(click slice)</span></div><canvas id="catShare"></canvas></div>
      <div class="chart-container"><div class="chart-title">Top SKUs in Selected Category <span class="chart-hint">(click bar)</span></div><canvas id="catTopSKU"></canvas></div>
    </div>
  `;
}

function renderCategory(){
  document.getElementById('catScopeLabel').textContent='Showing: '+scopeText();
  const cats=M.categories;
  const skusByCat={};
  cats.forEach(cat=>{skusByCat[cat]=filteredSKUs().filter(s=>s.cat===cat);});

  // Heatmap table with forecast
  let h='<table><thead><tr><th>Category</th><th class="r">'+periodLabel()+' $</th><th class="r">YTD $</th><th class="r">YTD LY</th><th class="r">YTD % Chg</th><th class="r">YTD Fcst</th><th class="r">vs Fcst</th><th class="r">Units</th></tr></thead><tbody>';
  const catMetrics=[];
  cats.forEach(cat=>{
    const skus=skusByCat[cat]||[];
    const prd=skus.reduce((a,s)=>a+periodVal(s).ty,0);
    const ytd=skus.reduce((a,s)=>a+s.ytd,0);
    const ytd_ly=skus.reduce((a,s)=>a+s.ytd_ly,0);
    const ytd_u=skus.reduce((a,s)=>a+s.ytd_u,0);
    const fc=skus.reduce((a,s)=>a+(s.fc_d_ytd||0),0);
    const pct=delta(ytd,ytd_ly);
    const fcPct=fc>0?delta(ytd,fc):null;
    catMetrics.push({cat,prd,ytd,ytd_ly,ytd_u,fc,pct,fcPct});
    const cls=pct==null?'hm-neutral':(pct>=.1?'hm-green':pct>=0?'hm-light-green':pct>=-0.1?'hm-light-red':'hm-red');
    const fcCls=fcPct==null?'':(fcPct>=0?'hm-green':'hm-red');
    h+=`<tr class="clickable" onclick="state.category='${cat}';document.getElementById('fCategory').value='${cat}';renderActive();">`;
    h+=`<td><strong>${cat}</strong></td><td class="r">${fmtK(prd)}</td><td class="r">${fmtK(ytd)}</td><td class="r">${fmtK(ytd_ly)}</td>`;
    h+=`<td class="r ${cls}" style="font-weight:700">${fmtPct(pct)}</td>`;
    h+=`<td class="r">${fc>0?fmtK(fc):'N/A'}</td><td class="r ${fcCls}" style="font-weight:700">${fcPct!=null?fmtPct(fcPct):'N/A'}</td>`;
    h+=`<td class="r">${fmtN(ytd_u)}</td></tr>`;
  });
  const tw=catMetrics.reduce((a,c)=>a+c.prd,0),ty=catMetrics.reduce((a,c)=>a+c.ytd,0),tly=catMetrics.reduce((a,c)=>a+c.ytd_ly,0),tu=catMetrics.reduce((a,c)=>a+c.ytd_u,0),tfc=catMetrics.reduce((a,c)=>a+c.fc,0);
  h+=`<tr style="font-weight:700;border-top:2px solid var(--border);"><td>TOTAL</td><td class="r">${fmtK(tw)}</td><td class="r">${fmtK(ty)}</td><td class="r">${fmtK(tly)}</td><td class="r">${fmtPct(delta(ty,tly))}</td><td class="r">${tfc>0?fmtK(tfc):'N/A'}</td><td class="r">${tfc>0?fmtPct(delta(ty,tfc)):'N/A'}</td><td class="r">${fmtN(tu)}</td></tr>`;
  h+='</tbody></table>';
  document.getElementById('catHeatmap').innerHTML=h;

  // Category monthly trend
  dc('catWeekly');
  const months=DATA.cat_monthly.filter(r=>r.yr===M.current_year).map(r=>r.mo);
  const uMos=[...new Set(months)].sort((a,b)=>(MO_ORDER[a]||0)-(MO_ORDER[b]||0));
  const datasets2=cats.map((cat,i)=>{
    const cm=DATA.cat_monthly.filter(r=>r.yr===M.current_year&&r.cat===cat);
    const cmMap=Object.fromEntries(cm.map(r=>[r.mo,r]));
    return{label:cat,data:uMos.map(m=>(cmMap[m]||{}).ty||0),borderColor:COLORS[i%COLORS.length],borderWidth:2,tension:.3,fill:false};
  });
  charts['catWeekly']=new Chart(document.getElementById('catWeekly'),{type:'line',data:{
    labels:uMos,datasets:datasets2
  },options:{responsive:true,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}}}});

  // Category share doughnut
  dc('catShare');
  const catYTD=catMetrics.map(c=>c.ytd);
  charts['catShare']=new Chart(document.getElementById('catShare'),{type:'doughnut',data:{
    labels:cats,datasets:[{data:catYTD,backgroundColor:COLORS.slice(0,cats.length)}]
  },options:{responsive:true,plugins:{legend:{position:'bottom'},tooltip:{callbacks:{label:ctx=>{const t=catYTD.reduce((a,b)=>a+b,0);return fmtK(ctx.raw)+' ('+(ctx.raw/(t||1)*100).toFixed(1)+'%)';}}}},
    onClick:(evt,elements)=>{
      if(!elements.length)return;
      const cat=cats[elements[0].index];
      const catSkus=filteredSKUs().filter(s=>s.cat===cat).sort((a,b)=>b.ytd-a.ytd);
      showDrillDown(cat+' SKU Detail',catSkus,getDDCols());
    }
  }});

  // Top SKUs in selected/first category
  dc('catTopSKU');
  const selCat=state.category!=='All'?state.category:cats[0];
  const topCatSKUs=DATA.skus.filter(s=>s.cat===selCat).slice(0,10);
  charts['catTopSKU']=new Chart(document.getElementById('catTopSKU'),{type:'bar',data:{
    labels:topCatSKUs.map(s=>s.desc.substring(0,25)),
    datasets:[
      {label:'TY YTD',data:topCatSKUs.map(s=>s.ytd),backgroundColor:'#C77D8A'},
      {label:'Forecast',data:topCatSKUs.map(s=>s.fc_d_ytd||0),backgroundColor:'rgba(91,155,213,.35)'},
    ]
  },options:{indexAxis:'y',responsive:true,plugins:{legend:{position:'top'}},scales:{x:{ticks:{callback:v=>fmtK(v)}}},
    onClick:(evt,elements)=>{if(elements.length)showSKUDetail(topCatSKUs[elements[0].index].ic);}
  }});
}

// =====================================================================
// TAB 5: FORECAST VS ACTUALS
// =====================================================================
function buildForecastPage(){
  document.getElementById('page-forecast').innerHTML=`
    <div class="scope-label" id="fcScopeLabel">Forecast scope: Sephora, Kohls</div>
    <div class="kpi-grid" id="fcKpis"></div>
    <div id="fcRiskCallouts"></div>
    <div class="chart-row full"><div class="chart-container" style="position:relative;"><div class="chart-title">Cumulative Pacing &mdash; Actual vs Forecast vs LY</div><div style="position:relative;height:300px;"><canvas id="fcPacing"></canvas></div></div></div>
    <div class="chart-row">
      <div class="chart-container"><div class="chart-title">Forecast Accuracy by Retailer (YTD)</div><canvas id="fcRetAcc"></canvas></div>
      <div class="chart-container"><div class="chart-title">Monthly Forecast vs Actual</div><canvas id="fcMonthly"></canvas></div>
    </div>
    <div class="table-container"><div class="table-title">Forecast Accuracy by Category</div><div id="fcCatTable"></div></div>
    <div class="table-container"><div class="table-title">SKU-Level Forecast Detail</div><div id="fcSKUTable"></div></div>
  `;
}

function renderForecast(){
  const fcRets=M.forecast_retailers||[];
  document.getElementById('fcScopeLabel').textContent='Forecast available for: '+(fcRets.length>0?fcRets.join(', '):'No forecast data');

  const rs=getAggSummary();
  const pv=periodVal(rs);

  // KPIs — now period-aware (fixes bug)
  const pvTy=pv.ty;
  const pvFc=pv.fc;
  const variance=pvTy-pvFc;
  const varPct=pvFc>0?variance/pvFc:null;
  const accRate=pvFc>0?(pvTy/pvFc*100).toFixed(1)+'%':'-';

  let kh='';
  kh+=`<div class="kpi-card"><div class="kpi-label">${periodLabel()} Actual $</div><div class="kpi-value">${fmtK(pvTy)}</div>${deltaHtml(pvTy,pv.ly,'vs LY')}</div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">${periodLabel()} Forecast $</div><div class="kpi-value">${pvFc>0?fmtK(pvFc):'N/A'}</div><div class="kpi-sub">Demand Plan</div></div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">${periodLabel()} Variance</div><div class="kpi-value ${pvFc>0?(variance>=0?'pos':'neg'):''}">${pvFc>0?fmtK(variance):'-'}</div>${pvFc>0?'<span class="kpi-delta '+(variance>=0?'pos':'neg')+'">'+fmtPct(varPct)+' vs Plan</span>':''}</div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">Accuracy Rate</div><div class="kpi-value">${accRate}</div><div class="kpi-sub">Actual / Forecast</div></div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">YTD Actual $</div><div class="kpi-value">${fmtK(rs.ytd)}</div><div class="kpi-sub">All periods</div></div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">YTD Forecast $</div><div class="kpi-value">${rs.fc_d_ytd>0?fmtK(rs.fc_d_ytd):'N/A'}</div><div class="kpi-sub">Through W${M.current_week}</div></div>`;
  document.getElementById('fcKpis').innerHTML=kh;

  // --- Risk callouts ---
  const atRisk=DATA.skus.filter(s=>s.risk==='critical'||s.risk==='warning').sort((a,b)=>{
    const aVar=a.fc_d_ytd>0?a.ytd-a.fc_d_ytd:-Infinity;
    const bVar=b.fc_d_ytd>0?b.ytd-b.fc_d_ytd:-Infinity;
    return aVar-bVar; // most negative first
  }).slice(0,12);

  if(atRisk.length>0){
    let rh=`<div class="risk-section"><div class="risk-header">&#9888; ${atRisk.length} SKUs Significantly Below Plan</div><div class="risk-grid">`;
    atRisk.forEach(s=>{
      const cls=s.risk==='critical'?'risk-critical':'risk-warning';
      rh+=`<div class="risk-item ${cls}" onclick="showSKUDetail(${s.ic})">`;
      rh+=`<span class="r-badge">${s.risk==='critical'?'CRITICAL':'WARNING'}</span>`;
      rh+=`<span class="r-name">${s.desc.substring(0,30)}</span>`;
      rh+=`<span class="r-var">${s.risk_pct!=null?s.risk_pct+'%':''} of plan</span>`;
      rh+=`</div>`;
    });
    rh+=`</div></div>`;
    document.getElementById('fcRiskCallouts').innerHTML=rh;
  } else {
    document.getElementById('fcRiskCallouts').innerHTML='';
  }

  // --- Cumulative pacing ---
  dc('fcPacing');
  const er=effRet()||'All';
  const wkTY=DATA.weekly.filter(r=>r.yr===M.current_year&&r.ret===er).sort((a,b)=>a.wk-b.wk);
  const wkLY=DATA.weekly.filter(r=>r.yr===M.current_year-1&&r.ret===er).sort((a,b)=>a.wk-b.wk);

  const tyMap2=Object.fromEntries(wkTY.map(r=>[r.wk,r]));
  const lyMap2=Object.fromEntries(wkLY.map(r=>[r.wk,r]));

  // Determine week range to show based on period
  let showWeeks;
  if(state.period==='mtd'){
    showWeeks=M.l4w_weeks.length>0?Array.from({length:M.current_week},(_,i)=>i+1):Array.from({length:M.current_week},(_,i)=>i+1);
  } else {
    const maxWk=Math.max(M.current_week,52);
    showWeeks=Array.from({length:maxWk},(_,i)=>i+1);
  }

  let cty=0,cly=0,cfc=0;
  const cumTY=[],cumLY=[],cumFC=[];
  showWeeks.forEach(w=>{
    const wkData=tyMap2[w];
    cty+=(wkData||{}).ty||0;cumTY.push(w<=M.current_week?cty:null);
    cly+=(lyMap2[w]||{}).ty||0;cumLY.push(lyMap2[w]?cly:null);
    cfc+=(wkData||{}).fc_d||0;cumFC.push(cfc>0?cfc:null);
  });

  charts['fcPacing']=new Chart(document.getElementById('fcPacing'),{type:'line',data:{
    labels:showWeeks.map(w=>'W'+w),
    datasets:[
      {label:'Actual (TY)',data:cumTY,borderColor:'#C77D8A',borderWidth:2.5,tension:.3,pointRadius:0},
      {label:'Forecast',data:cumFC,borderColor:'#5B9BD5',borderDash:[6,3],borderWidth:2,tension:.3,pointRadius:0},
      {label:'LY Actual',data:cumLY,borderColor:'#A89889',borderDash:[3,3],borderWidth:1.5,tension:.3,pointRadius:0},
    ]
  },options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}}}});

  // --- Accuracy by retailer ---
  dc('fcRetAcc');
  const retAcc=DATA.ret_summary.filter(r=>r.ret!=='All'&&r.fc_d_ytd>0);
  const accLabels=retAcc.map(r=>r.ret);
  const accData=retAcc.map(r=>(r.ytd/r.fc_d_ytd)*100);
  const accColors=accData.map(v=>v>=90&&v<=110?'#5CB85C':v>=80&&v<=120?'#D4A017':'#D9534F');
  charts['fcRetAcc']=new Chart(document.getElementById('fcRetAcc'),{type:'bar',data:{
    labels:accLabels,datasets:[{data:accData,backgroundColor:accColors}]
  },options:{indexAxis:'y',responsive:true,plugins:{legend:{display:false},tooltip:{callbacks:{label:ctx=>ctx.raw.toFixed(1)+'% of forecast'}}},scales:{x:{min:0,max:200,ticks:{callback:v=>v+'%'}}}}});

  // --- Monthly forecast vs actual ---
  dc('fcMonthly');
  const moData=DATA.monthly.filter(r=>r.yr===M.current_year&&r.ret===er).sort((a,b)=>a.mo_n-b.mo_n);
  charts['fcMonthly']=new Chart(document.getElementById('fcMonthly'),{type:'bar',data:{
    labels:moData.map(r=>r.mo),
    datasets:[
      {label:'Actual $',data:moData.map(r=>r.ty),backgroundColor:'#C77D8A'},
      {label:'Forecast $',data:moData.map(r=>r.fc_d),backgroundColor:'rgba(91,155,213,.5)'},
    ]
  },options:{responsive:true,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}}}});

  // --- Category accuracy table ---
  const catAccData=M.categories.map(cat=>{
    const skus=filteredSKUs().filter(s=>s.cat===cat);
    const act=skus.reduce((a,s)=>a+s.ytd,0);
    const fc=skus.reduce((a,s)=>a+(s.fc_d_ytd||0),0);
    return{cat,act,fc,acc:fc>0?(act/fc*100):null};
  });
  let ct='<table><thead><tr><th>Category</th><th class="r">YTD Actual $</th><th class="r">YTD Forecast $</th><th class="r">Accuracy %</th><th class="r">Variance $</th></tr></thead><tbody>';
  catAccData.forEach(c=>{
    const cls=c.acc==null?'hm-neutral':(c.acc>=90&&c.acc<=110?'hm-green':c.acc>=80&&c.acc<=120?'hm-light-green':'hm-red');
    ct+=`<tr><td><strong>${c.cat}</strong></td><td class="r">${fmtK(c.act)}</td><td class="r">${c.fc>0?fmtK(c.fc):'N/A'}</td><td class="r ${cls}" style="font-weight:700">${c.acc!=null?c.acc.toFixed(1)+'%':'-'}</td><td class="r">${c.fc>0?fmtK(c.act-c.fc):'-'}</td></tr>`;
  });
  ct+='</tbody></table>';
  document.getElementById('fcCatTable').innerHTML=ct;

  // --- SKU forecast table ---
  const fcSKUs=filteredSKUs().filter(s=>s.fc_d_ytd>0||s.ytd>0).sort((a,b)=>Math.abs(b.ytd-(b.fc_d_ytd||0))-Math.abs(a.ytd-(a.fc_d_ytd||0))).slice(0,50);
  let st='<table id="fcSKUTbl"><thead><tr>';
  const fcCols=[{l:'SKU',t:'str'},{l:'Description',t:'str'},{l:'Category',t:'str'},{l:'N/C',t:'str'},{l:'YTD Actual $',t:'num',a:'r'},{l:'YTD Fcst $',t:'num',a:'r'},{l:'Accuracy %',t:'num',a:'r'},{l:'Variance $',t:'num',a:'r'},{l:'Risk',t:'str'}];
  fcCols.forEach((c,i)=>{st+=`<th class="${c.a||''}" onclick="sortTable('fcSKUTbl',${i},'${c.t}')">${c.l}<span class="sort-arrow">&#9650;</span></th>`;});
  st+='</tr></thead><tbody>';
  fcSKUs.forEach(s=>{
    const acc=s.fc_d_ytd>0?(s.ytd/s.fc_d_ytd*100):null;
    const cls=acc==null?'':(acc>=90&&acc<=110?'pos':acc>=80&&acc<=120?'':'neg');
    const vari=s.ytd-(s.fc_d_ytd||0);
    const riskBadge=s.risk==='critical'?'<span class="badge badge-red">CRITICAL</span>':s.risk==='warning'?'<span class="badge badge-yellow">WARNING</span>':s.risk==='over'?'<span class="badge badge-blue">OVER</span>':s.risk==='on_track'?'<span class="badge badge-green">ON TRACK</span>':'-';
    st+=`<tr class="clickable" onclick="showSKUDetail(${s.ic})">`;
    st+=`<td>${s.ic}</td><td>${s.desc.substring(0,30)}</td><td>${s.cat}</td><td>${s.nvc}</td>`;
    st+=`<td class="r" data-v="${s.ytd}">${fmtK(s.ytd)}</td>`;
    st+=`<td class="r" data-v="${s.fc_d_ytd}">${s.fc_d_ytd>0?fmtK(s.fc_d_ytd):'N/A'}</td>`;
    st+=`<td class="r ${cls}" data-v="${acc||0}">${acc!=null?acc.toFixed(1)+'%':'-'}</td>`;
    st+=`<td class="r" data-v="${vari}">${fmtK(vari)}</td>`;
    st+=`<td>${riskBadge}</td>`;
    st+='</tr>';
  });
  st+='</tbody></table>';
  document.getElementById('fcSKUTable').innerHTML=st;
}

// =====================================================================
// TAB 6: NEW LAUNCHES
// =====================================================================
function buildNewLaunchesPage(){
  document.getElementById('page-launches').innerHTML=`
    <div class="scope-label" id="nlScopeLabel">New SKUs: Items with "New v Core = New" or no LY sales</div>
    <div class="kpi-grid" id="nlKpis"></div>
    <div class="chart-row full"><div class="chart-container" style="position:relative;"><div class="chart-title">New SKU Weekly Revenue vs Forecast</div><div style="position:relative;height:300px;"><canvas id="nlWeekly"></canvas></div></div></div>
    <div class="chart-row">
      <div class="chart-container"><div class="chart-title">Top 10 New SKUs <span class="chart-hint">(click bar for detail)</span></div><canvas id="nlTop10"></canvas></div>
      <div class="chart-container"><div class="chart-title">New SKUs by Category</div><canvas id="nlByCat"></canvas></div>
    </div>
    <div class="table-container"><div class="table-title">All New SKUs</div><div id="nlTableWrap"></div></div>
  `;
}

function renderNewLaunches(){
  const NL=DATA.new_launches;
  const newSkus=DATA.skus.filter(s=>s.is_new);
  const totalRevYTD=newSkus.reduce((a,s)=>a+s.ytd,0);
  const totalFcYTD=newSkus.reduce((a,s)=>a+(s.fc_d_ytd||0),0);
  const allYTD=DATA.skus.reduce((a,s)=>a+s.ytd,0);
  const pctOfTotal=allYTD>0?(totalRevYTD/allYTD*100).toFixed(1)+'%':'-';
  const avgPerSKU=newSkus.length>0?totalRevYTD/newSkus.length:0;
  const fcAcc=totalFcYTD>0?(totalRevYTD/totalFcYTD*100).toFixed(1)+'%':'N/A';

  let kh='';
  kh+=`<div class="kpi-card"><div class="kpi-label">New SKU Count</div><div class="kpi-value">${NL.count}</div><div class="kpi-sub">${NL.with_sales} with sales, ${NL.zero_sales} zero</div></div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">New SKU Revenue YTD</div><div class="kpi-value">${fmtK(totalRevYTD)}</div><div class="kpi-sub">${pctOfTotal} of total sales</div></div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">% of Total Sales</div><div class="kpi-value">${pctOfTotal}</div><div class="kpi-sub">New SKUs contribution</div></div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">Avg $ / New SKU</div><div class="kpi-value">${fmtK(avgPerSKU)}</div><div class="kpi-sub">${NL.count} SKUs</div></div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">vs Forecast Accuracy</div><div class="kpi-value">${fcAcc}</div><div class="kpi-sub">${fmtK(totalRevYTD)} actual / ${totalFcYTD>0?fmtK(totalFcYTD):'N/A'} plan</div></div>`;
  kh+=`<div class="kpi-card"><div class="kpi-label">Zero-Sales SKUs</div><div class="kpi-value" style="color:${NL.zero_sales>0?'var(--red)':'var(--green)'}">${NL.zero_sales}</div><div class="kpi-sub">No YTD revenue</div></div>`;
  document.getElementById('nlKpis').innerHTML=kh;

  // Weekly trend
  dc('nlWeekly');
  const aggWk=NL.agg_weekly||[];
  const wks=aggWk.map(w=>w.wk);
  const dsNL=[
    {label:'TY Revenue',data:aggWk.map(w=>w.ty),borderColor:'#C77D8A',backgroundColor:'rgba(199,125,138,.1)',borderWidth:2,fill:true,tension:.3},
  ];
  if(aggWk.some(w=>w.fc_d>0)){
    dsNL.push({label:'Forecast',data:aggWk.map(w=>w.fc_d),borderColor:'#5B9BD5',borderDash:[4,4],borderWidth:1.5,tension:.3,fill:false});
  }
  if(aggWk.some(w=>w.ly>0)){
    dsNL.push({label:'LY Revenue',data:aggWk.map(w=>w.ly),borderColor:'#A89889',borderDash:[5,3],borderWidth:1.5,tension:.3,fill:false});
  }
  charts['nlWeekly']=new Chart(document.getElementById('nlWeekly'),{type:'line',data:{
    labels:wks.map(w=>'W'+w),datasets:dsNL
  },options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}}}});

  // Top 10 new SKUs
  dc('nlTop10');
  const topNew=newSkus.filter(s=>s.ytd>0).sort((a,b)=>b.ytd-a.ytd).slice(0,10);
  const nlDS=[{label:'TY YTD',data:topNew.map(s=>s.ytd),backgroundColor:'#C77D8A'}];
  if(topNew.some(s=>s.fc_d_ytd>0)){
    nlDS.push({label:'Forecast YTD',data:topNew.map(s=>s.fc_d_ytd||0),backgroundColor:'rgba(91,155,213,.35)'});
  }
  charts['nlTop10']=new Chart(document.getElementById('nlTop10'),{type:'bar',data:{
    labels:topNew.map(s=>s.desc.substring(0,25)),datasets:nlDS
  },options:{indexAxis:'y',responsive:true,plugins:{legend:{position:'top'}},scales:{x:{ticks:{callback:v=>fmtK(v)}}},
    onClick:(evt,elements)=>{if(elements.length)showSKUDetail(topNew[elements[0].index].ic);}
  }});

  // By category
  dc('nlByCat');
  const catData={};
  newSkus.forEach(s=>{if(!catData[s.cat])catData[s.cat]={cat:s.cat,ty:0,fc:0,count:0};catData[s.cat].ty+=s.ytd;catData[s.cat].fc+=s.fc_d_ytd||0;catData[s.cat].count++;});
  const catArr=Object.values(catData).sort((a,b)=>b.ty-a.ty);
  charts['nlByCat']=new Chart(document.getElementById('nlByCat'),{type:'bar',data:{
    labels:catArr.map(c=>c.cat),
    datasets:[
      {label:'TY Revenue',data:catArr.map(c=>c.ty),backgroundColor:'#C77D8A'},
      {label:'Forecast',data:catArr.map(c=>c.fc),backgroundColor:'rgba(91,155,213,.35)'},
    ]
  },options:{responsive:true,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}}}});

  // Table
  let th='<table id="nlTable"><thead><tr>';
  const nlCols=[{l:'#',t:'num',a:'r'},{l:'Item',t:'num'},{l:'Description',t:'str'},{l:'Category',t:'str'},{l:'Collection',t:'str'},{l:'Launch',t:'str'},{l:'YTD $',t:'num',a:'r'},{l:'Fcst $',t:'num',a:'r'},{l:'Accuracy',t:'num',a:'r'},{l:'Risk',t:'str'}];
  nlCols.forEach((c,i)=>{th+=`<th class="${c.a||''}" onclick="sortTable('nlTable',${i},'${c.t}')">${c.l}<span class="sort-arrow">&#9650;</span></th>`;});
  th+='</tr></thead><tbody>';
  newSkus.sort((a,b)=>b.ytd-a.ytd).forEach((s,i)=>{
    const acc=s.fc_d_ytd>0?(s.ytd/s.fc_d_ytd*100):null;
    const accCls=acc==null?'':(acc>=90&&acc<=110?'pos':acc>=80&&acc<=120?'':'neg');
    const riskBadge=s.risk==='critical'?'<span class="badge badge-red">CRIT</span>':s.risk==='warning'?'<span class="badge badge-yellow">WARN</span>':s.risk==='on_track'?'<span class="badge badge-green">OK</span>':'-';
    th+=`<tr class="clickable" onclick="showSKUDetail(${s.ic})">`;
    th+=`<td class="r" data-v="${i+1}">${i+1}</td>`;
    th+=`<td data-v="${s.ic}">${s.ic}</td>`;
    th+=`<td>${s.desc.substring(0,30)}</td>`;
    th+=`<td>${s.cat}</td><td>${s.coll.substring(0,15)}</td><td>${s.ls||'-'}</td>`;
    th+=`<td class="r" data-v="${s.ytd}">${fmtK(s.ytd)}</td>`;
    th+=`<td class="r" data-v="${s.fc_d_ytd}">${s.fc_d_ytd>0?fmtK(s.fc_d_ytd):'N/A'}</td>`;
    th+=`<td class="r ${accCls}" data-v="${acc||0}">${acc!=null?acc.toFixed(1)+'%':'-'}</td>`;
    th+=`<td>${riskBadge}</td>`;
    th+='</tr>';
  });
  th+='</tbody></table>';
  document.getElementById('nlTableWrap').innerHTML=th;
}

// =====================================================================
// TAB 7: SEPHORA DOORS
// =====================================================================
let _doorState={territory:'All',region:'All',state:'All',fixture:'All',volume:'All',search:'',regionTile:null};

function buildDoorsPage(){
  const lm=DATA.loc_meta||{};
  const hasDoors=DATA.loc_sales&&DATA.loc_sales.length>0;
  if(!hasDoors){
    document.getElementById('page-doors').innerHTML='<div style="padding:40px;text-align:center;color:var(--text-muted);"><h3>Sephora Door Data Not Available</h3><p>Add the Sephora Sales Location Database file to the data folder and rebuild.</p></div>';
    return;
  }
  let fh='<div class="door-filters">';
  fh+='<div><label>Territory</label><select id="dFTerritory" onchange="_doorState.territory=this.value;_doorState.regionTile=null;renderDoors()"><option value="All">All Territories</option>';
  (lm.territories||[]).forEach(t=>{fh+='<option value="'+t+'">'+t+'</option>';});
  fh+='</select></div>';
  fh+='<div><label>Region</label><select id="dFRegion" onchange="_doorState.region=this.value;_doorState.regionTile=null;renderDoors()"><option value="All">All Regions</option>';
  (lm.regions||[]).forEach(t=>{fh+='<option value="'+t+'">'+t+'</option>';});
  fh+='</select></div>';
  fh+='<div><label>State</label><select id="dFState" onchange="_doorState.state=this.value;renderDoors()"><option value="All">All States</option>';
  (lm.states||[]).forEach(t=>{fh+='<option value="'+t+'">'+t+'</option>';});
  fh+='</select></div>';
  fh+='<div><label>Fixture</label><select id="dFFixture" onchange="_doorState.fixture=this.value;renderDoors()"><option value="All">All Fixtures</option>';
  (lm.fixtures||[]).forEach(t=>{fh+='<option value="'+t+'">'+t+'</option>';});
  fh+='</select></div>';
  fh+='<div><label>Volume</label><select id="dFVolume" onchange="_doorState.volume=this.value;renderDoors()"><option value="All">All Volumes</option>';
  (lm.volumes||[]).forEach(t=>{fh+='<option value="'+t+'">'+t+'</option>';});
  fh+='</select></div>';
  fh+='<div><label>Search</label><input class="door-search" id="dFSearch" placeholder="Store name or #..." oninput="_doorState.search=this.value.toLowerCase();renderDoors()"></div>';
  fh+='</div>';

  document.getElementById('page-doors').innerHTML=fh+`
    <div class="kpi-grid" id="doorKpis"></div>
    <div id="doorRegionGrid"></div>
    <div class="chart-row full"><div class="chart-container" style="position:relative;"><div class="chart-title">Weekly Sales Trend &mdash; Filtered Doors (TY vs LY)</div><div style="position:relative;height:280px;"><canvas id="doorWeekly"></canvas></div></div></div>
    <div class="table-container"><div class="table-title" style="display:flex;justify-content:space-between;align-items:center;"><span>Sephora Door Leaderboard</span><span id="doorCount" style="font-size:11px;color:var(--text-muted);font-weight:400;"></span></div><div id="doorTableWrap" style="max-height:600px;overflow-y:auto;"></div></div>
  `;
}

function filteredDoors(){
  let d=DATA.loc_sales||[];
  if(_doorState.territory!=='All')d=d.filter(r=>r.terr===_doorState.territory);
  if(_doorState.region!=='All')d=d.filter(r=>r.reg===_doorState.region);
  if(_doorState.regionTile)d=d.filter(r=>r.reg===_doorState.regionTile);
  if(_doorState.state!=='All')d=d.filter(r=>r.st===_doorState.state);
  if(_doorState.fixture!=='All')d=d.filter(r=>r.fix===_doorState.fixture);
  if(_doorState.volume!=='All')d=d.filter(r=>r.vol===_doorState.volume);
  if(_doorState.search){
    const q=_doorState.search;
    d=d.filter(r=>r.name.toLowerCase().includes(q)||String(r.loc).includes(q)||r.city.toLowerCase().includes(q));
  }
  return d;
}

function renderDoors(){
  const doors=filteredDoors();
  const totalWk=doors.reduce((a,d)=>a+d.wk,0);
  const totalYTD=doors.reduce((a,d)=>a+d.ytd,0);
  const totalYTDLY=doors.reduce((a,d)=>a+d.ytd_ly,0);
  const avgPerDoor=doors.length>0?totalYTD/doors.length:0;
  const topDoor=doors.length>0?doors[0]:null;

  // KPIs
  let kh='';
  kh+='<div class="kpi-card"><div class="kpi-label"># Doors</div><div class="kpi-value">'+fmtN(doors.length)+'</div><div class="kpi-sub">Filtered Sephora locations</div></div>';
  kh+='<div class="kpi-card"><div class="kpi-label">Week '+M.current_week+' Sales</div><div class="kpi-value">'+fmtK(totalWk)+'</div><div class="kpi-sub">All filtered doors</div></div>';
  kh+='<div class="kpi-card"><div class="kpi-label">YTD Sales</div><div class="kpi-value">'+fmtK(totalYTD)+'</div>'+deltaHtml(totalYTD,totalYTDLY,'vs LY')+'</div>';
  kh+='<div class="kpi-card"><div class="kpi-label">Avg $/Door YTD</div><div class="kpi-value">'+fmtK(avgPerDoor)+'</div><div class="kpi-sub">'+fmtN(doors.length)+' doors</div></div>';
  kh+='<div class="kpi-card"><div class="kpi-label">Top Door</div><div class="kpi-value" style="font-size:16px;">'+(topDoor?topDoor.name.substring(0,22):'—')+'</div><div class="kpi-sub">'+(topDoor?fmtK(topDoor.ytd)+' YTD':'')+'</div></div>';
  document.getElementById('doorKpis').innerHTML=kh;

  // Region heatmap tiles
  const regData={};
  doors.forEach(d=>{
    if(!d.reg||d.reg==='nan')return;
    if(!regData[d.reg])regData[d.reg]={reg:d.reg,ytd:0,ytd_ly:0,count:0,wk:0};
    regData[d.reg].ytd+=d.ytd;regData[d.reg].ytd_ly+=d.ytd_ly;regData[d.reg].count++;regData[d.reg].wk+=d.wk;
  });
  const regArr=Object.values(regData).sort((a,b)=>b.ytd-a.ytd);
  if(regArr.length>0&&regArr.length<=30){
    let rh='<div class="region-grid">';
    regArr.forEach(r=>{
      const d=r.ytd_ly>0?delta(r.ytd,r.ytd_ly):null;
      const cls=d==null?'':(d>=0?'pos':'neg');
      const active=_doorState.regionTile===r.reg?' active-filter':'';
      rh+='<div class="region-tile'+active+'" onclick="toggleRegionTile(\''+r.reg.replace(/'/g,"\\'")+'\')">';
      rh+='<div class="rt-label">'+r.reg+'</div>';
      rh+='<div class="rt-value">'+fmtK(r.ytd)+'</div>';
      rh+='<div class="rt-sub">'+r.count+' doors &middot; <span class="'+cls+'">'+fmtPct(d)+'</span></div>';
      rh+='</div>';
    });
    rh+='</div>';
    document.getElementById('doorRegionGrid').innerHTML=rh;
  } else {
    document.getElementById('doorRegionGrid').innerHTML='';
  }

  // Weekly trend for filtered doors
  dc('doorWeekly');
  const locWeeklyMap={};
  (DATA.loc_weekly||[]).forEach(lw=>{locWeeklyMap[lw.loc]=lw.weeks;});
  const filteredLocs=new Set(doors.map(d=>d.loc));
  const aggWk={};
  Object.entries(locWeeklyMap).forEach(([loc,weeks])=>{
    if(!filteredLocs.has(Number(loc)))return;
    weeks.forEach(w=>{
      if(!aggWk[w.wk])aggWk[w.wk]={wk:w.wk,ty:0,ly:0};
      aggWk[w.wk].ty+=w.ty;aggWk[w.wk].ly+=w.ly;
    });
  });
  const wkArr=Object.values(aggWk).sort((a,b)=>a.wk-b.wk);
  if(wkArr.length>0){
    charts['doorWeekly']=new Chart(document.getElementById('doorWeekly'),{type:'line',data:{
      labels:wkArr.map(w=>'W'+w.wk),
      datasets:[
        {label:'TY',data:wkArr.map(w=>w.ty),borderColor:'#C77D8A',backgroundColor:'rgba(199,125,138,.1)',borderWidth:2,fill:true,tension:.3},
        {label:'LY',data:wkArr.map(w=>w.ly),borderColor:'#A89889',borderDash:[5,3],borderWidth:1.5,tension:.3,fill:false},
      ]
    },options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmtK(v)}}}}});
  }

  // Door leaderboard table
  document.getElementById('doorCount').textContent=doors.length+' doors';
  let th='<table id="doorTable"><thead><tr>';
  const dCols=[
    {l:'#',t:'num',a:'r'},{l:'Location',t:'str'},{l:'City',t:'str'},{l:'Region',t:'str'},
    {l:'Vol',t:'str'},{l:'Fixture',t:'str'},{l:'Week $',t:'num',a:'r'},{l:'YTD $',t:'num',a:'r'},
    {l:'YTD Avg/Wk',t:'num',a:'r'},{l:'L5W Avg',t:'num',a:'r'},{l:'Chg vs Avg',t:'num',a:'r'},
    {l:'Chg %',t:'num',a:'r'},{l:'vs LY %',t:'num',a:'r'},
  ];
  dCols.forEach((c,i)=>{th+='<th class="'+(c.a||'')+'" onclick="sortTable(\'doorTable\','+i+',\''+c.t+'\')">'+c.l+'<span class="sort-arrow">&#9650;</span></th>';});
  th+='</tr></thead><tbody>';
  doors.forEach((d,i)=>{
    const lyPct=d.ytd_ly>0?delta(d.ytd,d.ytd_ly):null;
    const lyCls=lyPct==null?'':(lyPct>=0?'pos':'neg');
    const chgCls=d.chg_d>=0?'pos':'neg';
    th+='<tr class="clickable" onclick="showDoorDetail('+d.loc+')">';
    th+='<td class="r" data-v="'+(i+1)+'">'+(i+1)+'</td>';
    th+='<td data-v="'+d.name+'"><strong>'+d.name.substring(0,28)+'</strong></td>';
    th+='<td>'+d.city+'</td><td>'+d.reg+'</td>';
    th+='<td>'+d.vol+'</td><td>'+(d.fix.length>15?d.fix.substring(0,15)+'…':d.fix)+'</td>';
    th+='<td class="r" data-v="'+d.wk+'">'+fmt(d.wk)+'</td>';
    th+='<td class="r" data-v="'+d.ytd+'">'+fmtK(d.ytd)+'</td>';
    th+='<td class="r" data-v="'+d.ytd_avg+'">'+fmt(d.ytd_avg)+'</td>';
    th+='<td class="r" data-v="'+d.l5w_avg+'">'+fmt(d.l5w_avg)+'</td>';
    th+='<td class="r '+chgCls+'" data-v="'+d.chg_d+'">'+fmt(d.chg_d)+'</td>';
    th+='<td class="r '+chgCls+'" data-v="'+(d.chg_p||0)+'">'+fmtPct(d.chg_p)+'</td>';
    th+='<td class="r '+lyCls+'" data-v="'+(lyPct||0)+'">'+fmtPct(lyPct)+'</td>';
    th+='</tr>';
  });
  th+='</tbody></table>';
  document.getElementById('doorTableWrap').innerHTML=th;
}

function toggleRegionTile(reg){
  if(_doorState.regionTile===reg){_doorState.regionTile=null;}
  else{_doorState.regionTile=reg;}
  renderDoors();
}

function showDoorDetail(loc){
  const door=DATA.loc_sales.find(d=>d.loc===loc);
  if(!door)return;
  const locWeekly=(DATA.loc_weekly||[]).find(lw=>lw.loc===loc);
  let h='';
  if(locWeekly&&locWeekly.weeks.length>0){
    h+='<div class="modal-chart"><canvas id="doorModalChart"></canvas></div>';
  }
  h+='<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px;margin:12px 0 16px;">';
  h+='<div class="kpi-card" style="padding:12px;cursor:default;"><div class="kpi-label">Week $ (W'+M.current_week+')</div><div class="kpi-value" style="font-size:18px;">'+fmt(door.wk)+'</div></div>';
  h+='<div class="kpi-card" style="padding:12px;cursor:default;"><div class="kpi-label">YTD $</div><div class="kpi-value" style="font-size:18px;">'+fmtK(door.ytd)+'</div>'+deltaHtml(door.ytd,door.ytd_ly,'vs LY')+'</div>';
  h+='<div class="kpi-card" style="padding:12px;cursor:default;"><div class="kpi-label">YTD Avg/Wk</div><div class="kpi-value" style="font-size:18px;">'+fmt(door.ytd_avg)+'</div></div>';
  h+='<div class="kpi-card" style="padding:12px;cursor:default;"><div class="kpi-label">L5W Avg</div><div class="kpi-value" style="font-size:18px;">'+fmt(door.l5w_avg)+'</div></div>';
  h+='<div class="kpi-card" style="padding:12px;cursor:default;"><div class="kpi-label">Week Units</div><div class="kpi-value" style="font-size:18px;">'+fmtN(door.wk_u)+'</div></div>';
  h+='<div class="kpi-card" style="padding:12px;cursor:default;"><div class="kpi-label">Inventory $</div><div class="kpi-value" style="font-size:18px;">'+fmtK(door.inv)+'</div><div class="kpi-sub">AFS: '+fmtN(door.afs)+' units</div></div>';
  h+='</div>';
  h+='<div style="margin-top:12px;font-size:12px;color:var(--text-muted);">';
  h+='<strong>Territory:</strong> '+door.terr+' &middot; <strong>Region:</strong> '+door.reg+' &middot; <strong>State:</strong> '+door.st+' &middot; <strong>City:</strong> '+door.city;
  h+='<br><strong>Fixture:</strong> '+door.fix+' &middot; <strong>Store Volume:</strong> '+door.vol;
  h+='</div>';
  showModal(door.name+' (#'+door.loc+')',h);
  document.getElementById('modalCount').textContent=door.terr+' | '+door.reg+' | '+door.vol;
  if(locWeekly&&locWeekly.weeks.length>0){
    dc('doorModalChart');
    const wks=locWeekly.weeks.sort((a,b)=>a.wk-b.wk);
    charts['doorModalChart']=new Chart(document.getElementById('doorModalChart'),{type:'line',data:{
      labels:wks.map(w=>'W'+w.wk),
      datasets:[
        {label:'TY $',data:wks.map(w=>w.ty),borderColor:'#C77D8A',backgroundColor:'rgba(199,125,138,.1)',borderWidth:2,fill:true,tension:.3},
        {label:'LY $',data:wks.map(w=>w.ly),borderColor:'#A89889',borderDash:[5,3],borderWidth:1.5,tension:.3,fill:false},
      ]
    },options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top'}},scales:{y:{ticks:{callback:v=>fmt(v)}}}}});
  }
}

// =====================================================================
// INIT
// =====================================================================
buildOverviewPage();
buildSKUPage();
buildRetailerPage();
buildCategoryPage();
buildForecastPage();
buildNewLaunchesPage();
buildDoorsPage();
buildRetailerMultiSelect();
renderOverview();

</script>
</body>
</html>'''

# ---------------------------------------------------------------------------
# Write output
# ---------------------------------------------------------------------------
html = HTML_TEMPLATE
html = html.replace("%%DATA_PLACEHOLDER%%", json_data)
html = html.replace("%%FONT_CSS%%", FONT_CSS)

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    f.write(html)

file_size_mb = os.path.getsize(OUTPUT_FILE) / (1024 * 1024)
print(f"        Output: {OUTPUT_FILE}")
print(f"        File size: {file_size_mb:.1f} MB")
print()
print("=" * 60)
print(f"  Dashboard built successfully!")
print(f"  Data as of: FY{CURRENT_YEAR} Week {CURRENT_WEEK} ({CURRENT_MONTH})")
print(f"  SKUs: {len(sku_perf_list)}")
print(f"  Retailers: {len(all_retailers)}")
print(f"  Forecast data points: {total_fcst_rows:,}")
if fcst_warnings:
    print(f"  Warnings: {len(fcst_warnings)}")
    for w in fcst_warnings:
        print(f"    - {w}")
print("=" * 60)
